"""Daily EscortsAndBabes Lookup sync into safety-screening watchlist."""

from __future__ import annotations

import logging
import os
import re
import time
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import config
from core.settings_manager import get_setting, set_setting
from services.safety_screening_service import replace_watchlist
from utils.phone_normalization import extract_normalized_au_mobile

logger = logging.getLogger("adella_chatbot.ugly_mugs_sync")

BASE_URL = "https://www.escortsandbabes.com.au"
LOGIN_URL = f"{BASE_URL}/Login/"
LOOKUP_PAGE_URL_TEMPLATE = BASE_URL + "/Lookup?page={page}"

DEFAULT_TOTAL_PAGES = 14935
DEFAULT_START_PAGE = 1
DEFAULT_SCHEDULE_HOUR = 9
DEFAULT_SCHEDULE_MINUTE = 0
DEFAULT_SCHEDULE_TIMEZONE = "Australia/Adelaide"
DEFAULT_PAGE_DELAY_SECONDS = 0.12

LOGIN_USERNAME_FIELD = (
    "p$lt$ctl10$pageplaceholder$p$lt$ctl00$WebPartZone$WebPartZone_zone$"
    "EBLogonForm$Login1$UserName"
)
LOGIN_PASSWORD_FIELD = (
    "p$lt$ctl10$pageplaceholder$p$lt$ctl00$WebPartZone$WebPartZone_zone$"
    "EBLogonForm$Login1$Password"
)
LOGIN_BUTTON_FIELD = (
    "p$lt$ctl10$pageplaceholder$p$lt$ctl00$WebPartZone$WebPartZone_zone$"
    "EBLogonForm$Login1$LoginButton"
)

MAX_PAGE_FETCH_ATTEMPTS = 5
REQUEST_TIMEOUT_SECONDS = 45
MOBILE_RE = re.compile(
    r"(?<!\d)(?:\+?61[\s-]?4\d{2}[\s-]?\d{3}[\s-]?\d{3}|0?4\d{2}[\s-]?\d{3}[\s-]?\d{3})(?!\d)"
)

SETTING_SYNC_ENABLED = "safety_screening_ugly_mugs_sync_enabled"
SETTING_USERNAME = "safety_screening_ugly_mugs_username"
SETTING_PASSWORD = "safety_screening_ugly_mugs_password"
SETTING_START_PAGE = "safety_screening_ugly_mugs_start_page"
SETTING_TOTAL_PAGES = "safety_screening_ugly_mugs_total_pages"
SETTING_EXPORT_PATH = "safety_screening_ugly_mugs_export_path"
SETTING_PAGE_DELAY_SECONDS = "safety_screening_ugly_mugs_page_delay_seconds"
SETTING_SCHEDULE_HOUR = "safety_screening_ugly_mugs_sync_hour"
SETTING_SCHEDULE_MINUTE = "safety_screening_ugly_mugs_sync_minute"
SETTING_SCHEDULE_TIMEZONE = "safety_screening_ugly_mugs_sync_timezone"

LEGACY_SETTING_SYNC_ENABLED = "ugly_mugs_sync_enabled"
LEGACY_SETTING_USERNAME = "ugly_mugs_lookup_username"
LEGACY_SETTING_PASSWORD = "ugly_mugs_lookup_password"
LEGACY_SETTING_TOTAL_PAGES = "ugly_mugs_lookup_total_pages"
LEGACY_SETTING_EXPORT_PATH = "ugly_mugs_sync_output_xlsx"
LEGACY_SETTING_SCHEDULE_HOUR = "ugly_mugs_sync_hour_local"
LEGACY_SETTING_SCHEDULE_MINUTE = "ugly_mugs_sync_minute_local"
LEGACY_SETTING_SCHEDULE_TIMEZONE = "ugly_mugs_sync_timezone"

META_LAST_RUN_AT = "safety_screening_ugly_mugs_last_run_at"
META_LAST_STATUS = "safety_screening_ugly_mugs_last_status"
META_LAST_COUNT = "safety_screening_ugly_mugs_last_count"
META_LAST_ERROR = "safety_screening_ugly_mugs_last_error"
META_LAST_DURATION = "safety_screening_ugly_mugs_last_duration_seconds"
META_LAST_EXPORT_PATH = "safety_screening_ugly_mugs_last_export_path"
META_LAST_FAILED_PAGES = "safety_screening_ugly_mugs_last_failed_pages"
META_LAST_UNIQUE_COUNT = "safety_screening_ugly_mugs_last_unique_count"
META_LAST_PAGE_RANGE = "safety_screening_ugly_mugs_last_page_range"
META_LAST_STARTED_AT = "safety_screening_ugly_mugs_last_started_at"

LEGACY_META_LAST_RUN_AT = "ugly_mugs_sync_last_run_at"
LEGACY_META_LAST_STATUS = "ugly_mugs_sync_last_status"
LEGACY_META_LAST_COUNT = "ugly_mugs_sync_last_count"
LEGACY_META_LAST_ERROR = "ugly_mugs_sync_last_error"

MIRRORED_SETTING_KEYS: dict[str, tuple[str, ...]] = {
    META_LAST_RUN_AT: (LEGACY_META_LAST_RUN_AT,),
    META_LAST_STATUS: (LEGACY_META_LAST_STATUS,),
    META_LAST_COUNT: (LEGACY_META_LAST_COUNT,),
    META_LAST_ERROR: (LEGACY_META_LAST_ERROR,),
}


@dataclass(frozen=True)
class UglyMugsSyncConfig:
    enabled: bool
    username: str
    password: str
    start_page: int
    total_pages: int
    export_path: str
    page_delay_seconds: float


@dataclass(frozen=True)
class UglyMugsSyncSchedule:
    hour: int
    minute: int
    timezone: str
    enabled: bool


def _as_bool(value: object, default: bool = False) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "on"}


def _as_setting_keys(setting_keys: str | tuple[str, ...]) -> tuple[str, ...]:
    return (setting_keys,) if isinstance(setting_keys, str) else setting_keys


def _read_setting_text(setting_keys: str | tuple[str, ...], env_keys: tuple[str, ...], default: str = "") -> str:
    for setting_key in _as_setting_keys(setting_keys):
        db_val = str(get_setting(setting_key) or "").strip()
        if db_val:
            return db_val
    for env_key in env_keys:
        env_val = str(os.environ.get(env_key, "")).strip()
        if env_val:
            return env_val
    return default


def _read_setting_int(
    setting_keys: str | tuple[str, ...],
    env_keys: tuple[str, ...],
    default: int,
    *,
    minimum: int,
    maximum: int,
) -> int:
    raw = _read_setting_text(setting_keys, env_keys, default=str(default))
    try:
        value = int(str(raw).strip())
    except (TypeError, ValueError):
        logger.warning("Invalid int setting for %s=%r; using %s", setting_keys, raw, default)
        value = default
    return max(minimum, min(maximum, value))


def _read_setting_float(
    setting_keys: str | tuple[str, ...],
    env_keys: tuple[str, ...],
    default: float,
    *,
    minimum: float,
    maximum: float,
) -> float:
    raw = _read_setting_text(setting_keys, env_keys, default=str(default))
    try:
        value = float(str(raw).strip())
    except (TypeError, ValueError):
        logger.warning("Invalid float setting for %s=%r; using %s", setting_keys, raw, default)
        value = default
    return max(minimum, min(maximum, value))


def _read_setting_bool(
    setting_keys: str | tuple[str, ...], env_keys: tuple[str, ...], default: bool = False
) -> bool:
    raw = _read_setting_text(setting_keys, env_keys, default="")
    if not raw:
        return default
    return _as_bool(raw, default=default)


def _safe_set_setting(key: str, value: str) -> None:
    for target_key in (key, *MIRRORED_SETTING_KEYS.get(key, ())):
        ok = set_setting(target_key, value)
        if not ok:
            logger.warning("Failed persisting ugly-mugs sync metadata: %s", target_key)


def _default_export_path() -> str:
    base_dir = Path(getattr(config, "BASE_DIR", os.getcwd())).resolve()
    return str(base_dir / "ugly_mugs_lookup_watchlist.xlsx")


def _resolve_export_path(raw_path: str) -> str:
    expanded = os.path.expanduser((raw_path or "").strip())
    if not expanded:
        return _default_export_path()
    if os.path.isabs(expanded):
        return expanded
    return str(Path(getattr(config, "BASE_DIR", os.getcwd())).resolve() / expanded)


def get_ugly_mugs_sync_config() -> UglyMugsSyncConfig:
    enabled = _read_setting_bool(
        (SETTING_SYNC_ENABLED, LEGACY_SETTING_SYNC_ENABLED),
        ("UGLY_MUGS_SYNC_ENABLED",),
        default=False,
    )
    username = _read_setting_text(
        (SETTING_USERNAME, LEGACY_SETTING_USERNAME),
        ("UGLY_MUGS_LOOKUP_USERNAME", "UGLY_MUGS_USERNAME", "ESCORTS_BABES_USER"),
        default="",
    )
    password = _read_setting_text(
        (SETTING_PASSWORD, LEGACY_SETTING_PASSWORD),
        ("UGLY_MUGS_LOOKUP_PASSWORD", "UGLY_MUGS_PASSWORD", "ESCORTS_BABES_PASS"),
        default="",
    )
    start_page = _read_setting_int(
        SETTING_START_PAGE,
        ("UGLY_MUGS_START_PAGE",),
        default=DEFAULT_START_PAGE,
        minimum=1,
        maximum=10_000_000,
    )
    total_pages = _read_setting_int(
        (SETTING_TOTAL_PAGES, LEGACY_SETTING_TOTAL_PAGES),
        ("UGLY_MUGS_LOOKUP_TOTAL_PAGES", "UGLY_MUGS_TOTAL_PAGES"),
        default=DEFAULT_TOTAL_PAGES,
        minimum=1,
        maximum=10_000_000,
    )
    if total_pages < start_page:
        logger.warning(
            "Configured ugly-mugs page range invalid (start=%s total=%s); clamping total=start",
            start_page,
            total_pages,
        )
        total_pages = start_page

    export_path = _resolve_export_path(
        _read_setting_text(
            (SETTING_EXPORT_PATH, LEGACY_SETTING_EXPORT_PATH),
            ("UGLY_MUGS_SYNC_OUTPUT_XLSX", "UGLY_MUGS_EXPORT_PATH"),
            default=_default_export_path(),
        )
    )
    page_delay_seconds = _read_setting_float(
        SETTING_PAGE_DELAY_SECONDS,
        ("UGLY_MUGS_PAGE_DELAY_SECONDS",),
        default=DEFAULT_PAGE_DELAY_SECONDS,
        minimum=0.0,
        maximum=5.0,
    )
    return UglyMugsSyncConfig(
        enabled=enabled,
        username=username,
        password=password,
        start_page=start_page,
        total_pages=total_pages,
        export_path=export_path,
        page_delay_seconds=page_delay_seconds,
    )


def get_ugly_mugs_sync_schedule() -> UglyMugsSyncSchedule:
    cfg = get_ugly_mugs_sync_config()
    hour = _read_setting_int(
        (SETTING_SCHEDULE_HOUR, LEGACY_SETTING_SCHEDULE_HOUR),
        ("UGLY_MUGS_SYNC_HOUR_LOCAL", "UGLY_MUGS_SYNC_HOUR"),
        default=DEFAULT_SCHEDULE_HOUR,
        minimum=0,
        maximum=23,
    )
    minute = _read_setting_int(
        (SETTING_SCHEDULE_MINUTE, LEGACY_SETTING_SCHEDULE_MINUTE),
        ("UGLY_MUGS_SYNC_MINUTE_LOCAL", "UGLY_MUGS_SYNC_MINUTE"),
        default=DEFAULT_SCHEDULE_MINUTE,
        minimum=0,
        maximum=59,
    )
    timezone_name = _read_setting_text(
        (SETTING_SCHEDULE_TIMEZONE, LEGACY_SETTING_SCHEDULE_TIMEZONE),
        ("UGLY_MUGS_SYNC_TIMEZONE",),
        default=DEFAULT_SCHEDULE_TIMEZONE,
    )
    return UglyMugsSyncSchedule(hour=hour, minute=minute, timezone=timezone_name, enabled=cfg.enabled)


def _write_metadata(
    *,
    started_at: datetime,
    completed_at: datetime,
    status: str,
    inserted_count: int,
    unique_count: int,
    failed_pages_count: int,
    export_path: str,
    error: str = "",
    page_range: str = "",
) -> None:
    _safe_set_setting(META_LAST_STARTED_AT, started_at.isoformat())
    _safe_set_setting(META_LAST_RUN_AT, completed_at.isoformat())
    _safe_set_setting(META_LAST_STATUS, status)
    _safe_set_setting(META_LAST_COUNT, str(max(0, int(inserted_count))))
    _safe_set_setting(META_LAST_UNIQUE_COUNT, str(max(0, int(unique_count))))
    _safe_set_setting(META_LAST_FAILED_PAGES, str(max(0, int(failed_pages_count))))
    _safe_set_setting(META_LAST_DURATION, str(max(0, int((completed_at - started_at).total_seconds()))))
    _safe_set_setting(META_LAST_EXPORT_PATH, export_path.strip())
    _safe_set_setting(META_LAST_PAGE_RANGE, page_range.strip())
    _safe_set_setting(META_LAST_ERROR, (error or "").strip()[:1000])


def _create_scraper():
    try:
        import cloudscraper
    except ImportError as exc:
        raise RuntimeError("cloudscraper dependency is required for ugly-mugs sync") from exc
    return cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )


def _parse_html(html: str):
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:
        raise RuntimeError("beautifulsoup4 dependency is required for ugly-mugs sync") from exc
    return BeautifulSoup(html or "", "html.parser")


def _is_login_page(html: str) -> bool:
    text = (html or "").lower()
    return "eblogonform_login1_username" in text or ("/login/" in text and "sign in" in text)


def _login(scraper: Any, username: str, password: str) -> None:
    response = scraper.get(LOGIN_URL, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    soup = _parse_html(response.text)
    form = soup.find("form")
    if form is None:
        raise RuntimeError("Lookup login form not found")

    payload = {
        i.get("name"): i.get("value", "")
        for i in form.find_all("input")
        if i.get("name")
    }
    payload[LOGIN_USERNAME_FIELD] = username
    payload[LOGIN_PASSWORD_FIELD] = password
    payload[LOGIN_BUTTON_FIELD] = "Sign in"

    post_response = scraper.post(
        LOGIN_URL,
        data=payload,
        timeout=REQUEST_TIMEOUT_SECONDS,
        allow_redirects=True,
    )
    post_response.raise_for_status()
    if _is_login_page(post_response.text):
        raise RuntimeError("Lookup login failed. Check ugly-mugs credentials.")


def _extract_lookup_numbers(html: str) -> list[str]:
    candidates = MOBILE_RE.findall(html or "")
    soup = _parse_html(html or "")
    for anchor in soup.select('a[href^="tel:"]'):
        anchor_get = getattr(anchor, "get", None)
        if not callable(anchor_get):
            continue
        tel = str(anchor_get("href") or "").replace("tel:", "").strip()
        if tel:
            candidates.append(tel)
    return candidates


def _scrape_lookup_numbers(
    scraper: Any,
    *,
    username: str,
    password: str,
    start_page: int,
    total_pages: int,
    page_delay_seconds: float,
) -> tuple[dict[str, dict[str, Any]], list[tuple[int, str]]]:
    unique_numbers: dict[str, dict[str, Any]] = {}
    failed_pages: list[tuple[int, str]] = []

    for page in range(start_page, total_pages + 1):
        page_url = LOOKUP_PAGE_URL_TEMPLATE.format(page=page)
        html = ""
        last_error: Exception | None = None
        for attempt in range(1, MAX_PAGE_FETCH_ATTEMPTS + 1):
            try:
                response = scraper.get(page_url, timeout=REQUEST_TIMEOUT_SECONDS)
                if response.status_code != 200:
                    raise RuntimeError(f"status={response.status_code}")
                html = response.text or ""
                if _is_login_page(html):
                    logger.warning("Lookup session expired on page %s; re-authenticating", page)
                    _login(scraper, username, password)
                    raise RuntimeError("session expired; relogged")
                last_error = None
                break
            except Exception as exc:
                last_error = exc
                if attempt >= MAX_PAGE_FETCH_ATTEMPTS:
                    break
                time.sleep(min(2**attempt, 8))

        if last_error is not None:
            msg = str(last_error)
            failed_pages.append((page, msg[:500]))
            logger.error("Lookup page failed (%s): %s", page, msg)
            continue

        seen_page: set[str] = set()
        for raw in _extract_lookup_numbers(html):
            normalized = extract_normalized_au_mobile(raw)
            if not normalized or normalized in seen_page:
                continue
            seen_page.add(normalized)
            if normalized not in unique_numbers:
                unique_numbers[normalized] = {
                    "raw": str(raw or "").strip(),
                    "first_page": page,
                    "occurrences": 1,
                }
            else:
                unique_numbers[normalized]["occurrences"] = int(unique_numbers[normalized]["occurrences"]) + 1

        if page_delay_seconds > 0:
            time.sleep(page_delay_seconds)

    return unique_numbers, failed_pages


def _write_export_workbook(
    *,
    export_path: str,
    unique_numbers: dict[str, dict[str, Any]],
    failed_pages: list[tuple[int, str]],
    started_at: datetime,
    completed_at: datetime,
    start_page: int,
    total_pages: int,
) -> str:
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.warning("openpyxl not installed; writing ugly-mugs export as CSV instead")
        export_abs_path = os.path.abspath(export_path)
        csv_path = os.path.splitext(export_abs_path)[0] + ".csv"
        csv_dir = os.path.dirname(csv_path)
        if csv_dir:
            os.makedirs(csv_dir, exist_ok=True)
        with open(csv_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Normalized Mobile", "Raw First Seen", "First Seen Page", "Occurrences"])
            for normalized in sorted(unique_numbers.keys()):
                meta = unique_numbers[normalized]
                writer.writerow(
                    [
                        normalized,
                        str(meta.get("raw", "")),
                        int(meta.get("first_page", 0)),
                        int(meta.get("occurrences", 0)),
                    ]
                )
        return csv_path

    export_abs_path = os.path.abspath(export_path)
    export_dir = os.path.dirname(export_abs_path)
    if export_dir:
        os.makedirs(export_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Lookup Mobiles"
    ws.append(["Normalized Mobile", "Raw First Seen", "First Seen Page", "Occurrences"])
    for normalized in sorted(unique_numbers.keys()):
        meta = unique_numbers[normalized]
        ws.append(
            [
                normalized,
                str(meta.get("raw", "")),
                int(meta.get("first_page", 0)),
                int(meta.get("occurrences", 0)),
            ]
        )

    failed_sheet = wb.create_sheet("Failed Pages")
    failed_sheet.append(["Page", "Error"])
    for page, message in failed_pages:
        failed_sheet.append([page, (message or "")[:300]])

    meta_sheet = wb.create_sheet("Meta")
    meta_sheet.append(["Key", "Value"])
    meta_sheet.append(["started_at_utc", started_at.isoformat()])
    meta_sheet.append(["completed_at_utc", completed_at.isoformat()])
    meta_sheet.append(["start_page", start_page])
    meta_sheet.append(["total_pages", total_pages])
    meta_sheet.append(["unique_count", len(unique_numbers)])
    meta_sheet.append(["failed_pages_count", len(failed_pages)])
    meta_sheet.append(["elapsed_seconds", int((completed_at - started_at).total_seconds())])

    tmp_path = f"{export_abs_path}.tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, export_abs_path)
    return export_abs_path


def run_ugly_mugs_sync() -> dict[str, Any]:
    """Run one full sync from EscortsAndBabes Lookup into safety-screening watchlist."""
    started_at = datetime.now(timezone.utc)
    cfg = get_ugly_mugs_sync_config()
    page_range = f"{cfg.start_page}-{cfg.total_pages}"

    if not cfg.enabled:
        completed_at = datetime.now(timezone.utc)
        _write_metadata(
            started_at=started_at,
            completed_at=completed_at,
            status="skipped",
            inserted_count=0,
            unique_count=0,
            failed_pages_count=0,
            export_path=cfg.export_path,
            error="Sync disabled",
            page_range=page_range,
        )
        logger.info("Ugly-mugs sync skipped: feature disabled")
        return {"status": "skipped", "reason": "disabled", "inserted": 0}

    if not cfg.username or not cfg.password:
        completed_at = datetime.now(timezone.utc)
        _write_metadata(
            started_at=started_at,
            completed_at=completed_at,
            status="skipped",
            inserted_count=0,
            unique_count=0,
            failed_pages_count=0,
            export_path=cfg.export_path,
            error="Missing credentials",
            page_range=page_range,
        )
        logger.warning("Ugly-mugs sync skipped: credentials missing")
        return {"status": "skipped", "reason": "missing_credentials", "inserted": 0}

    _safe_set_setting(META_LAST_STARTED_AT, started_at.isoformat())
    _safe_set_setting(META_LAST_STATUS, "running")
    _safe_set_setting(META_LAST_ERROR, "")

    try:
        scraper = _create_scraper()
        _login(scraper, cfg.username, cfg.password)
        unique_numbers, failed_pages = _scrape_lookup_numbers(
            scraper,
            username=cfg.username,
            password=cfg.password,
            start_page=cfg.start_page,
            total_pages=cfg.total_pages,
            page_delay_seconds=cfg.page_delay_seconds,
        )
        if not unique_numbers:
            raise RuntimeError("No valid AU mobile numbers were scraped from Lookup pages.")

        export_completed_at = datetime.now(timezone.utc)
        written_export_path = _write_export_workbook(
            export_path=cfg.export_path,
            unique_numbers=unique_numbers,
            failed_pages=failed_pages,
            started_at=started_at,
            completed_at=export_completed_at,
            start_page=cfg.start_page,
            total_pages=cfg.total_pages,
        )
        if not written_export_path:
            written_export_path = cfg.export_path

        watchlist_payload = [
            (normalized, str(meta.get("raw", "")).strip())
            for normalized, meta in sorted(unique_numbers.items(), key=lambda item: item[0])
        ]
        watchlist_result = replace_watchlist(
            watchlist_payload,
            filename=os.path.basename(written_export_path),
        )
        inserted = int(watchlist_result.get("inserted") or 0)
        completed_at = datetime.now(timezone.utc)

        _write_metadata(
            started_at=started_at,
            completed_at=completed_at,
            status="success",
            inserted_count=inserted,
            unique_count=len(unique_numbers),
            failed_pages_count=len(failed_pages),
            export_path=written_export_path,
            page_range=page_range,
        )
        logger.info(
            "Ugly-mugs sync complete: inserted=%s unique=%s failed_pages=%s export=%s",
            inserted,
            len(unique_numbers),
            len(failed_pages),
            cfg.export_path,
        )
        return {
            "status": "success",
            "inserted": inserted,
            "unique_count": len(unique_numbers),
            "failed_pages_count": len(failed_pages),
            "export_path": written_export_path,
            "page_range": page_range,
        }
    except Exception as exc:
        completed_at = datetime.now(timezone.utc)
        _write_metadata(
            started_at=started_at,
            completed_at=completed_at,
            status="failed",
            inserted_count=0,
            unique_count=0,
            failed_pages_count=0,
            export_path=cfg.export_path,
            error=str(exc),
            page_range=page_range,
        )
        logger.exception("Ugly-mugs sync failed: %s", exc)
        raise
