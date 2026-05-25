"""Config management blueprint - /config route."""

from utils.log_sanitize import LOG_SUPPRESSED_FMT


import json
import logging
import os
import io
import zipfile
from typing import Callable, cast
import defusedxml.ElementTree as ET

from flask import Blueprint, flash, redirect, render_template, request, session, url_for
import requests
from werkzeug.utils import secure_filename

from admin.auth import require_auth, verify_password
from config import BASE_DIR, get_google_maps_browser_api_key, get_google_maps_server_api_key, get_opencage_api_key
from core.settings_manager import get_all_settings, get_setting, set_setting
from utils.secret_mask import mask_secret_value
from services.safety_screening_service import get_watchlist_stats, replace_watchlist
from utils.phone_normalization import extract_normalized_au_mobile
from utils.row_utils import row_get

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logger = logging.getLogger("escort_chatbot.admin.config")

_CONFIG_BASE_DIR = cast(str, BASE_DIR)
_get_google_maps_browser_api_key = cast(Callable[[], str | None], get_google_maps_browser_api_key)
_get_google_maps_server_api_key = cast(Callable[[], str | None], get_google_maps_server_api_key)
_get_opencage_api_key = cast(Callable[[], str | None], get_opencage_api_key)

config_bp = Blueprint('config', __name__, template_folder='../templates')

# --- File upload endpoint for deployment ---
from flask import jsonify

_UPLOAD_DIR = os.path.join(_CONFIG_BASE_DIR, "uploads")
os.makedirs(_UPLOAD_DIR, exist_ok=True)

@config_bp.route('/upload', methods=['POST'])
def upload_file():
    if not (session.get("admin_authenticated") or session.get("config_authenticated")):
        return jsonify({"error": "Unauthorized"}), 401
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    # Validate file size (10MB limit)
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({"error": "File too large (max 10MB)"}), 400

    # Whitelist allowed extensions
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in {'.json', '.xlsx', '.xlsm', '.csv'}:
        return jsonify({"error": "Invalid file type"}), 400

    filename = secure_filename(file.filename or "")
    filename = os.path.basename(filename)
    save_path = os.path.join(_UPLOAD_DIR, filename)

    # Prevent overwriting existing files
    counter = 1
    base, file_ext = os.path.splitext(save_path)
    while os.path.exists(save_path):
        save_path = f"{base}_{counter}{file_ext}"
        counter += 1

    file.save(save_path)
    return jsonify({"success": True, "filename": os.path.basename(save_path)})
# --- End file upload endpoint ---
CREDENTIALS_JSON_PATH = os.path.join(_CONFIG_BASE_DIR, "credentials.json")

ESCORT_SMS_CATEGORIES = [
    ('deposit_validation_failed',  'Deposit validation failed'),
    ('outcall_notifications',      'Outcall notifications'),
    ('enquiry_forwarding',            'Enquiry forwarding (ENQUIRY <question>)'),
    ('refund_forwarding',            'Refund forwarding'),
    ('doubles_source_escort',      'Doubles booking — source escort alert (MFF/MMF)'),
    ('safety_screening',            'Safety screening match alerts'),
    ('special_bookings',            'Extended experience enquiry alerts (Overnight / Dirty Weekend / Fly Me To You) — escort notified immediately to personally contact client'),
    ('client_rating',                'Post-booking client rating (SMS with link to feedback webform)'),
    ('feedback_replies',            'Post-booking feedback — SMS replies to escort (3 STAR / N Y N acknowledgements)'),
    ('incall_client_forwards',      'Incall — all texts to your phone (1h before start)'),
    ('deposit_followup',            'Deposit follow-up reminder (sent to escort when client hasn\'t paid deposit after 4h)'),
    ('prebooking_checkin',          'Pre-booking check-in (sent to escort ~2h before a confirmed booking starts)'),
]

# Historic settings key (per-category escort SMS toggles)


def _is_config_authenticated():
    """Check if user is authenticated for config access."""
    return session.get("admin_authenticated", False) or session.get("config_authenticated", False)


def _is_admin_or_config_authenticated():
    """Check if user is authenticated for either admin or config access."""
    return (session.get("admin_authenticated", False)
            or session.get("config_authenticated", False)
            or session.get("database_authenticated", False))


def _bool_from_setting(setting_key: str, default: bool = True) -> bool:
    """Parse boolean from ``admin_settings`` only."""
    raw = get_setting(setting_key)
    if raw is None or str(raw).strip() == "":
        return default
    text = str(raw).strip().lower()
    return text in ("1", "true", "yes", "on")


def _int_from_setting(
    setting_key: str,
    default: int,
    min_value: int,
    max_value: int,
) -> int:
    """Parse integer from ``admin_settings`` only, clamped to range."""
    raw = get_setting(setting_key)
    if raw is None or str(raw).strip() == "":
        return default
    try:
        value = int(str(raw).strip())
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, value))


def _float_from_setting(
    setting_key: str,
    default: float,
    min_value: float,
    max_value: float,
) -> float:
    """Parse float from ``admin_settings`` only, clamped to range."""
    raw = get_setting(setting_key)
    if raw is None or str(raw).strip() == "":
        return default
    try:
        value = float(str(raw).strip())
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, value))


def _parse_optional_excel_int(value) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None

    text = str(value).strip().replace(',', '')
    if not text:
        return None
    try:
        if '.' in text:
            parsed = float(text)
            return int(parsed) if parsed.is_integer() else None
        return int(text)
    except (TypeError, ValueError):
        return None


def _xlsx_column_index(cell_ref: str) -> int | None:
    if not cell_ref:
        return None
    letters = []
    for ch in str(cell_ref):
        if ch.isalpha():
            letters.append(ch.upper())
        else:
            break
    if not letters:
        return None
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch) - ord('A') + 1)
    return index - 1


def _read_xlsx_shared_strings(archive: zipfile.ZipFile, names: set[str]) -> list[str]:
    if 'xl/sharedStrings.xml' not in names:
        return []

    shared_root = ET.fromstring(archive.read('xl/sharedStrings.xml'))
    return [
        ''.join((t.text or '') for t in si.findall('.//{*}t'))
        for si in shared_root.findall('.//{*}si')
    ]


def _read_xlsx_relationship_targets(archive: zipfile.ZipFile, names: set[str]) -> dict[str, str]:
    if 'xl/_rels/workbook.xml.rels' not in names:
        return {}

    rel_root = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
    rel_targets: dict[str, str] = {}
    for rel in rel_root.findall('.//{*}Relationship'):
        rel_id = rel.attrib.get('Id')
        target = rel.attrib.get('Target')
        if rel_id and target:
            rel_targets[rel_id] = target
    return rel_targets


def _normalize_xlsx_sheet_target(target: str) -> str:
    normalized_target = target.lstrip('/')
    if not normalized_target.startswith('xl/'):
        normalized_target = f'xl/{normalized_target}'
    return normalized_target


def _read_xlsx_sheet_paths(
    archive: zipfile.ZipFile,
    names: set[str],
    rel_targets: dict[str, str],
) -> list[str]:
    sheet_paths: list[str] = []
    if 'xl/workbook.xml' in names:
        workbook_root = ET.fromstring(archive.read('xl/workbook.xml'))
        rel_key = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
        for sheet in workbook_root.findall('.//{*}sheet'):
            rid = sheet.attrib.get(rel_key)
            target = rel_targets.get(rid or '')
            if target:
                sheet_paths.append(_normalize_xlsx_sheet_target(target))

    if sheet_paths:
        return sheet_paths
    if 'xl/worksheets/sheet1.xml' in names:
        return ['xl/worksheets/sheet1.xml']
    return []


def _extract_xlsx_cell_value(cell, shared_strings: list[str]) -> object:
    cell_type = (cell.attrib.get('t') or '').strip().lower()
    if cell_type == 'inlinestr':
        inline = cell.find('{*}is')
        if inline is None:
            return None
        return ''.join((t.text or '') for t in inline.findall('.//{*}t'))

    value_element = cell.find('{*}v')
    raw = (value_element.text or '').strip() if value_element is not None and value_element.text is not None else ''
    if not raw:
        return None
    if cell_type != 's':
        return raw

    try:
        shared_idx = int(raw)
    except (TypeError, ValueError):
        return raw
    if 0 <= shared_idx < len(shared_strings):
        return shared_strings[shared_idx]
    return raw


def _extract_xlsx_row_values(row, shared_strings: list[str]) -> dict[int, object]:
    values_by_col: dict[int, object] = {}
    for cell in row.findall('{*}c'):
        col_idx = _xlsx_column_index(cell.attrib.get('r', ''))
        if col_idx is not None:
            values_by_col[col_idx] = _extract_xlsx_cell_value(cell, shared_strings)
    return values_by_col


def _extract_xlsx_sheet_rows(
    archive: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: list[str],
) -> list[tuple[object, object, object]]:
    sheet_root = ET.fromstring(archive.read(sheet_path))
    rows: list[tuple[object, object, object]] = []
    for row in sheet_root.findall('.//{*}sheetData/{*}row'):
        values_by_col = _extract_xlsx_row_values(row, shared_strings)
        rows.append((values_by_col.get(0), values_by_col.get(2), values_by_col.get(3)))
    return rows


def _extract_watchlist_rows_from_xlsx_bytes(raw_bytes: bytes) -> list[tuple[object, object, object]]:
    _MAX_UNCOMPRESSED = 50 * 1024 * 1024  # 50MB
    _MAX_FILES = 1000

    rows: list[tuple[object, object, object]] = []
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
        total_size = sum(info.file_size for info in archive.infolist())
        if total_size > _MAX_UNCOMPRESSED:
            raise ValueError(f"XLSX file too large when uncompressed: {total_size} bytes")
        if len(archive.infolist()) > _MAX_FILES:
            raise ValueError(f"XLSX contains too many files: {len(archive.infolist())}")

        names = set(archive.namelist())
        shared_strings = _read_xlsx_shared_strings(archive, names)
        rel_targets = _read_xlsx_relationship_targets(archive, names)
        sheet_paths = _read_xlsx_sheet_paths(archive, names, rel_targets)
        for sheet_path in sheet_paths:
            if sheet_path in names:
                rows.extend(_extract_xlsx_sheet_rows(archive, sheet_path, shared_strings))
    return rows


def _handle_config_page_authentication() -> tuple[bool, str | None]:
    authenticated = _is_config_authenticated()
    error = None
    if request.method == 'POST' and not authenticated:
        password = request.form.get('password')
        if verify_password(password or ""):
            session["config_authenticated"] = True
            authenticated = True
        else:
            error = 'Invalid password'
    return authenticated, error


def _maybe_save_schedule_api_key() -> None:
    if request.method == 'POST' and 'schedule_api_key' in request.form:
        new_schedule_api_key = (request.form.get('schedule_api_key') or '').strip()
        if new_schedule_api_key:
            set_setting('schedule_api_key', new_schedule_api_key)


def _normalize_safety_screening_mode(mode: str) -> str:
    normalized_mode = (mode or 'warn_only').strip().lower()
    if normalized_mode not in ('warn_only', 'auto_block'):
        return 'warn_only'
    return normalized_mode


def _get_schedule_api_context() -> dict[str, object]:
    schedule_api_key_raw = get_setting('schedule_api_key') or ''
    return {
        'schedule_api_key_masked': mask_secret_value(schedule_api_key_raw) if schedule_api_key_raw else '',
    }


def _get_httpsms_watchlist_context() -> dict[str, object]:
    httpsms_api_key_raw = (get_setting('httpsms_api_key') or '').strip()
    httpsms_phone = (get_setting('httpsms_phone_number') or '').strip()
    safety_watchlist = get_watchlist_stats()
    return {
        'safety_screening_enabled': _bool_from_setting('safety_screening_enabled', default=False),
        'safety_screening_mode': _normalize_safety_screening_mode(get_setting('safety_screening_mode') or 'warn_only'),
        'safety_watchlist_count': safety_watchlist.get('count', 0),
        'safety_watchlist_last_uploaded_at': safety_watchlist.get('last_uploaded_at', ''),
        'safety_watchlist_last_uploaded_filename': safety_watchlist.get('last_uploaded_filename', ''),
        'httpsms_configured': bool(httpsms_api_key_raw and httpsms_phone),
        'httpsms_enabled': _bool_from_setting('httpsms_enabled', default=True),
        'httpsms_phone': httpsms_phone,
        'httpsms_api_key_set': bool(httpsms_api_key_raw),
        'httpsms_api_key_masked': mask_secret_value(httpsms_api_key_raw) if httpsms_api_key_raw else '',
    }


def _get_ai_keys_context() -> dict[str, object]:
    claude_key_raw = get_setting('claude_api_key') or ''
    gemini_key_raw = get_setting('gemini_api_key') or ''
    return {
        'claude_key': bool(claude_key_raw),
        'gemini_key': bool(gemini_key_raw),
        'claude_key_masked': mask_secret_value(claude_key_raw),
        'gemini_key_masked': mask_secret_value(gemini_key_raw),
    }


def _get_maps_and_opencage_context() -> dict[str, object]:
    maps_legacy_db_key = (get_setting('google_maps_api_key') or '').strip()
    return {
        'base_url': (get_setting('base_url') or '').strip().rstrip('/'),
        'google_maps_browser_key': bool((get_setting('google_maps_browser_api_key') or maps_legacy_db_key).strip()),
        'google_maps_server_key': bool((get_setting('google_maps_server_api_key') or maps_legacy_db_key).strip()),
        'google_maps_browser_key_masked': mask_secret_value(_get_google_maps_browser_api_key() or ''),
        'google_maps_server_key_masked': mask_secret_value(_get_google_maps_server_api_key() or ''),
        'opencage_key': bool((get_setting('opencage_api_key') or '').strip()),
        'opencage_key_masked': mask_secret_value(_get_opencage_api_key() or ''),
    }


def _get_sms_gateway_secret_context() -> dict[str, object]:
    sms_encryption_key_raw = (get_setting('sms_encryption_key') or '').strip()
    gateway_secret_raw = (get_setting('gateway_secret') or '').strip()
    webhook_signature_secret_raw = (get_setting('httpsms_webhook_signature_secret') or '').strip()
    signature_required_raw = (get_setting('httpsms_webhook_signature_required') or '').strip().lower()
    return {
        'sms_encryption_key_set': bool(sms_encryption_key_raw),
        'sms_encryption_key_masked': mask_secret_value(sms_encryption_key_raw) if sms_encryption_key_raw else '',
        'gateway_secret_set': bool(gateway_secret_raw),
        'gateway_secret_masked': mask_secret_value(gateway_secret_raw) if gateway_secret_raw else '',
        'webhook_signature_secret_set': bool(webhook_signature_secret_raw),
        'webhook_signature_secret_masked': mask_secret_value(webhook_signature_secret_raw) if webhook_signature_secret_raw else '',
        'httpsms_webhook_signature_required': (signature_required_raw not in ('false', '0', 'no')) if signature_required_raw else True,
    }


def _get_flask_secret_context() -> dict[str, object]:
    flask_secret_db = (get_setting('flask_secret_key') or '').strip()
    flask_secret_env = (os.environ.get('SECRET_KEY', '') or '').strip()
    return {
        'secret_key_set': bool(flask_secret_db or flask_secret_env),
        'flask_secret_key_masked': mask_secret_value(flask_secret_db) if flask_secret_db else (mask_secret_value(flask_secret_env) if flask_secret_env else ''),
    }


def _get_runtime_threshold_context() -> dict[str, object]:
    default_threshold = _float_from_setting(
        'ai_fallback_confidence_threshold',
        default=0.45,
        min_value=0.0,
        max_value=1.0,
    )
    return {
        'ai_fallback_confidence_threshold': round(default_threshold, 2),
        'ai_fallback_confidence_threshold_qualification': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_qualification',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
        'ai_fallback_confidence_threshold_availability': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_availability',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
        'ai_fallback_confidence_threshold_screening': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_screening',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
        'ai_fallback_confidence_threshold_deposit': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_deposit',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
        'ai_fallback_confidence_threshold_confirmation': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_confirmation',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
        'ai_fallback_confidence_threshold_follow_up': round(
            _float_from_setting(
                'ai_fallback_confidence_threshold_follow_up',
                default=default_threshold,
                min_value=0.0,
                max_value=1.0,
            ),
            2,
        ),
    }


def _get_runtime_config_context() -> dict[str, object]:
    flow_version_default = (get_setting('flow_version_default') or 'rollout').strip().lower()
    if flow_version_default not in ('rollout', 'v1', 'v2'):
        flow_version_default = 'rollout'
    return {
        'escort_phone': (get_setting('escort_phone_number') or '').strip(),
        'vision_configured': os.path.exists(CREDENTIALS_JSON_PATH),
        'database_url_set': bool(os.environ.get('DATABASE_URL', '').strip()),
        'run_startup_db_migrations': (get_setting('run_startup_db_migrations') or 'false').lower() in ('true', '1', 'yes'),
        'flow_version_default': flow_version_default,
        'flow_version_v2_rollout_percent': _int_from_setting(
            'flow_version_v2_rollout_percent',
            default=0,
            min_value=0,
            max_value=100,
        ),
    }


def _get_ugly_mugs_context() -> dict[str, object]:
    ugly_mugs_lookup_password_raw = get_setting('ugly_mugs_lookup_password') or ''
    return {
        'ugly_mugs_sync_enabled': _bool_from_setting('ugly_mugs_sync_enabled', default=False),
        'ugly_mugs_lookup_username': (get_setting('ugly_mugs_lookup_username') or '').strip(),
        'ugly_mugs_lookup_password_set': bool(ugly_mugs_lookup_password_raw.strip()),
        'ugly_mugs_lookup_password_masked': mask_secret_value(ugly_mugs_lookup_password_raw),
        'ugly_mugs_sync_hour_local': _int_from_setting(
            'ugly_mugs_sync_hour_local',
            default=9,
            min_value=0,
            max_value=23,
        ),
        'ugly_mugs_sync_minute_local': _int_from_setting(
            'ugly_mugs_sync_minute_local',
            default=0,
            min_value=0,
            max_value=59,
        ),
        'ugly_mugs_sync_timezone': (get_setting('ugly_mugs_sync_timezone') or 'Australia/Adelaide').strip() or 'Australia/Adelaide',
        'ugly_mugs_lookup_total_pages': _int_from_setting(
            'ugly_mugs_lookup_total_pages',
            default=14935,
            min_value=1,
            max_value=50000,
        ),
        'ugly_mugs_sync_output_xlsx': (get_setting('ugly_mugs_sync_output_xlsx') or '').strip(),
        'ugly_mugs_sync_last_run_at': (get_setting('ugly_mugs_sync_last_run_at') or '').strip(),
        'ugly_mugs_sync_last_status': (get_setting('ugly_mugs_sync_last_status') or '').strip(),
        'ugly_mugs_sync_last_count': (get_setting('ugly_mugs_sync_last_count') or '').strip(),
        'ugly_mugs_sync_last_error': (get_setting('ugly_mugs_sync_last_error') or '').strip(),
    }


def _get_redis_context() -> dict[str, object]:
    redis_env = os.environ.get('REDIS_URL')
    redis_db = get_setting('redis_url')
    redis_url_raw = (redis_env or redis_db or '').strip()
    if redis_env:
        redis_url_source = 'env (.env file)'
    elif redis_db:
        redis_url_source = 'database'
    else:
        redis_url_source = ''
    return {
        'redis_url_set': bool(redis_url_raw),
        'redis_url_masked': mask_secret_value(redis_url_raw) if redis_url_raw else '',
        'redis_url_source': redis_url_source,
    }


def _get_server_ip() -> str:
    try:
        return requests.get('https://api.ipify.org', timeout=3).text.strip()
    except Exception as e:
        logger.warning(LOG_SUPPRESSED_FMT, e)
        return ''


def _build_config_page_context() -> dict[str, object]:
    context: dict[str, object] = {'authenticated': True}
    for section in (
        _get_schedule_api_context(),
        _get_httpsms_watchlist_context(),
        _get_ai_keys_context(),
        _get_maps_and_opencage_context(),
        _get_sms_gateway_secret_context(),
        _get_flask_secret_context(),
        _get_runtime_config_context(),
        _get_runtime_threshold_context(),
        _get_ugly_mugs_context(),
        _get_redis_context(),
    ):
        context.update(section)
    context['server_ip'] = _get_server_ip()
    return context


@config_bp.route('/config', methods=['GET', 'POST'])
def config_page():
    """Configuration management page."""
    authenticated, error = _handle_config_page_authentication()
    if not authenticated:
        return render_template('config.html', authenticated=False, error=error)

    get_all_settings()
    _maybe_save_schedule_api_key()
    return render_template('config.html', **_build_config_page_context())


@config_bp.route('/config/save-ugly-mugs-sync', methods=['POST'])
def save_ugly_mugs_sync():
    """Save Ugly Mugs daily sync credentials and schedule settings."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    def _sanitize_int(raw: str, default: int, min_value: int, max_value: int) -> int:
        try:
            parsed = int((raw or '').strip())
        except (TypeError, ValueError):
            return default
        return max(min_value, min(max_value, parsed))

    enabled_present = request.form.get('enabled_present', '').strip() == '1'
    enabled = 'enabled' in request.form
    lookup_username = (request.form.get('lookup_username') or '').strip()[:255]
    lookup_password = (request.form.get('lookup_password') or '').strip()
    sync_hour_local = _sanitize_int(request.form.get('sync_hour_local', ''), default=9, min_value=0, max_value=23)
    sync_minute_local = _sanitize_int(request.form.get('sync_minute_local', ''), default=0, min_value=0, max_value=59)
    lookup_total_pages = _sanitize_int(
        request.form.get('lookup_total_pages', ''),
        default=14935,
        min_value=1,
        max_value=50000,
    )
    sync_timezone = ((request.form.get('sync_timezone') or '').strip() or 'Australia/Adelaide')[:128]
    sync_output_xlsx = (request.form.get('sync_output_xlsx') or '').strip()[:512]

    failed = []
    if enabled_present and not set_setting('ugly_mugs_sync_enabled', 'true' if enabled else 'false'):
        failed.append('sync enabled toggle')
    if not set_setting('ugly_mugs_lookup_username', lookup_username):
        failed.append('lookup username')
    if lookup_password and not set_setting('ugly_mugs_lookup_password', lookup_password):
        failed.append('lookup password')
    if not set_setting('ugly_mugs_sync_hour_local', str(sync_hour_local)):
        failed.append('sync hour')
    if not set_setting('ugly_mugs_sync_minute_local', str(sync_minute_local)):
        failed.append('sync minute')
    if not set_setting('ugly_mugs_sync_timezone', sync_timezone):
        failed.append('sync timezone')
    if not set_setting('ugly_mugs_lookup_total_pages', str(lookup_total_pages)):
        failed.append('total pages')
    if not set_setting('ugly_mugs_sync_output_xlsx', sync_output_xlsx):
        failed.append('output xlsx path')

    if failed:
        flash(
            'Some Ugly Mugs sync settings could not be saved: ' + ', '.join(failed) + '. '
            'Check server logs and DATABASE_URL.',
            'error',
        )
    else:
        flash('Ugly Mugs sync settings saved.', 'success')
        logger.info(
            "Ugly Mugs sync settings updated: enabled=%s, hour=%s, minute=%s, timezone=%s, pages=%s, output=%s",
            enabled,
            sync_hour_local,
            sync_minute_local,
            sync_timezone,
            lookup_total_pages,
            sync_output_xlsx or "(none)",
        )

    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-safety-screening', methods=['POST'])
def save_safety_screening():
    """Save safety-screening toggles and mode."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    enabled_present = request.form.get('enabled_present', '').strip() == '1'
    enabled = 'enabled' in request.form
    mode = (request.form.get('screening_mode') or 'warn_only').strip().lower()
    if mode not in ('warn_only', 'auto_block'):
        mode = 'warn_only'

    failed = []
    if enabled_present:
        if not set_setting('safety_screening_enabled', 'true' if enabled else 'false'):
            failed.append('screening enabled toggle')
    if not set_setting('safety_screening_mode', mode):
        failed.append('screening mode')

    if failed:
        flash(
            'Some safety-screening settings could not be saved: ' + ', '.join(failed) + '. '
            'Check server logs and DATABASE_URL.',
            'error',
        )
    else:
        logger.info("Safety screening settings updated: enabled=%s mode=%s", enabled, mode)
        flash('Safety screening settings saved.', 'success')
    return redirect(url_for('config.config_page'))


def _append_failed_setting(
    failed: list[str],
    label: str,
    setting_key: str,
    setting_value: object,
    should_save: bool = True,
) -> None:
    if should_save and not set_setting(setting_key, str(setting_value)):
        failed.append(label)


def _get_watchlist_upload_details():
    file = request.files.get('watchlist_file')
    if not file or not file.filename:
        flash('No watchlist file selected.', 'warning')
        return None, None

    # Validate file size before reading (20MB limit)
    file.stream.seek(0, os.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    if size > 20 * 1024 * 1024:
        flash('Watchlist file too large (max 20MB).', 'error')
        return None, None

    filename = secure_filename(file.filename or '')
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.xls':
        flash('Legacy .xls format is not supported here. Please save the file as .xlsx and upload again.', 'error')
        return None, None
    if ext not in ('.xlsx', '.xlsm'):
        flash('Watchlist file must be .xlsx or .xlsm.', 'error')
        return None, None
    return file, filename


def _extract_watchlist_rows_with_openpyxl(file) -> list[tuple[object, object, object]] | None:
    loader = load_workbook
    if loader is None:
        return None
    try:
        workbook = loader(file.stream, read_only=True, data_only=True)
        extracted_rows: list[tuple[object, object, object]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                extracted_rows.append((row_get(row, 0, None), row_get(row, 2, None), row_get(row, 3, None)))
        return extracted_rows
    except Exception as e:
        logger.error("Safety watchlist upload failed to read workbook via openpyxl: %s", e)
        flash('Could not read Excel file. Please check the file and try again.', 'error')
        return None


def _extract_watchlist_rows_with_fallback(file) -> list[tuple[object, object, object]] | None:
    try:
        file.stream.seek(0)
    except Exception as e:
        logger.warning("Safety watchlist upload: file.stream.seek(0) failed: %s", e)

    try:
        return _extract_watchlist_rows_from_xlsx_bytes(file.read())
    except ValueError as e:
        logger.error("Safety watchlist upload rejected: %s", e)
        flash(str(e), 'error')
        return None
    except Exception as e:
        logger.error("Safety watchlist upload fallback parser failed: %s", e)
        flash('Could not read Excel file on this server (missing parser dependency).', 'error')
        return None


def _extract_watchlist_rows_from_upload(file) -> list[tuple[object, object, object]] | None:
    if load_workbook is not None:
        return _extract_watchlist_rows_with_openpyxl(file)
    return _extract_watchlist_rows_with_fallback(file)


def _normalize_watchlist_numbers(
    extracted_rows: list[tuple[object, object, object]],
) -> list[tuple[str, str, int | None, int | None]]:
    numbers: list[tuple[str, str, int | None, int | None]] = []
    for raw_phone, recency_value, report_value in extracted_rows:
        normalized = extract_normalized_au_mobile(raw_phone)
        if not normalized:
            continue

        recency_rank = _parse_optional_excel_int(recency_value)
        if recency_rank is not None and recency_rank <= 0:
            recency_rank = None

        report_count = _parse_optional_excel_int(report_value)
        if report_count is None or report_count < 0:
            report_count = 0

        numbers.append((normalized, str(raw_phone or '').strip(), recency_rank, report_count))
    return numbers


def _replace_uploaded_watchlist(
    numbers: list[tuple[str, str, int | None, int | None]],
    filename: str,
) -> None:
    try:
        result = replace_watchlist(numbers, filename=filename)
        inserted = int(result.get('inserted') or 0)
        logger.info("Safety watchlist replaced from %s: %d numbers", filename, inserted)
        flash(f'Safety watchlist uploaded successfully: {inserted} numbers imported.', 'success')
    except Exception as e:
        logger.exception("Safety watchlist upload failed: %s", e)
        flash('Failed to import watchlist. Please check logs and database connectivity.', 'error')


@config_bp.route('/config/upload-safety-watchlist', methods=['POST'])
def upload_safety_watchlist():
    """Upload and replace safety-screening watchlist from Excel."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    file, filename = _get_watchlist_upload_details()
    if file is None or filename is None:
        return redirect(url_for('config.config_page'))

    extracted_rows = _extract_watchlist_rows_from_upload(file)
    if extracted_rows is None:
        return redirect(url_for('config.config_page'))

    numbers = _normalize_watchlist_numbers(extracted_rows)
    if not numbers:
        flash('No valid mobile numbers found in the uploaded spreadsheet.', 'warning')
        return redirect(url_for('config.config_page'))

    _replace_uploaded_watchlist(numbers, filename)
    return redirect(url_for('config.config_page'))


def _get_httpsms_form_data() -> dict[str, object]:
    return {
        'api_key': (request.form.get('httpsms_api_key') or '').strip(),
        'phone_number': request.form.get('phone_number', '').strip(),
        'enabled_present': request.form.get('enabled_present', '').strip() == '1',
        'enabled': 'enabled' in request.form,
        'sms_encryption_key': (request.form.get('sms_encryption_key') or '').strip(),
        'gateway_secret': (request.form.get('gateway_secret') or '').strip(),
        'webhook_signature_secret': (request.form.get('httpsms_webhook_signature_secret') or '').strip(),
        'sig_req_present': request.form.get('sig_req_present', '').strip() == '1',
        'sig_req_enabled': 'httpsms_webhook_signature_required' in request.form,
    }


def _has_httpsms_form_input(form_data: dict[str, object]) -> bool:
    return any(
        (
            form_data['api_key'],
            form_data['phone_number'],
            form_data['enabled_present'],
            form_data['sms_encryption_key'],
            form_data['gateway_secret'],
            form_data['webhook_signature_secret'],
            form_data['sig_req_present'],
        )
    )


def _save_httpsms_form_data(form_data: dict[str, object]) -> list[str]:
    failed: list[str] = []
    operations = (
        ('API key', 'httpsms_api_key', form_data['api_key'], bool(form_data['api_key'])),
        ('phone number', 'httpsms_phone_number', form_data['phone_number'], bool(form_data['phone_number'])),
        ('gateway enabled toggle', 'httpsms_enabled', 'true' if form_data['enabled'] else 'false', bool(form_data['enabled_present'])),
        ('sms encryption key', 'sms_encryption_key', form_data['sms_encryption_key'], bool(form_data['sms_encryption_key'])),
        ('gateway secret', 'gateway_secret', form_data['gateway_secret'], bool(form_data['gateway_secret'])),
        (
            'webhook signature secret',
            'httpsms_webhook_signature_secret',
            form_data['webhook_signature_secret'],
            bool(form_data['webhook_signature_secret']),
        ),
        (
            'signature required toggle',
            'httpsms_webhook_signature_required',
            'true' if form_data['sig_req_enabled'] else 'false',
            bool(form_data['sig_req_present']),
        ),
    )
    for label, setting_key, setting_value, should_save in operations:
        _append_failed_setting(failed, label, setting_key, setting_value, should_save)
    return failed


@config_bp.route('/config/save-httpsms', methods=['POST'])
def save_httpsms():
    """Save httpSMS gateway configuration."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    form_data = _get_httpsms_form_data()
    if not _has_httpsms_form_input(form_data):
        flash('No httpSMS settings provided.', 'warning')
        return redirect(url_for('config.config_page'))

    try:
        failed = _save_httpsms_form_data(form_data)
    except Exception as e:
        logger.exception("save_httpsms failed: %s", e)
        flash(
            f'Could not save httpSMS settings: {e!s}. '
            'Check server logs and that the database is reachable.',
            'error',
        )
        return redirect(url_for('config.config_page'))

    if failed:
        flash(
            'Some httpSMS settings could not be saved: ' + ', '.join(failed) + '. '
            'See server logs for details.',
            'error',
        )
    else:
        flash('httpSMS gateway settings saved successfully!', 'success')
        logger.info("httpSMS gateway configuration updated")

    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-ai-keys', methods=['POST'])
def save_ai_keys():
    """Save AI service API keys."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))
    claude_key = request.form.get('claude_key', '').strip()
    gemini_key = request.form.get('gemini_key', '').strip()

    failed: list[str] = []
    if claude_key:
        if set_setting('claude_api_key', claude_key):
            logger.info("Claude API key updated")
        else:
            failed.append('Claude API key')

    if gemini_key:
        if set_setting('gemini_api_key', gemini_key):
            logger.info("Gemini API key updated")
        else:
            failed.append('Gemini API key')

    if not claude_key and not gemini_key:
        flash('No keys provided', 'warning')
    elif failed:
        flash(
            'Could not save to database: ' + ', '.join(failed) + '. Check server logs and DATABASE_URL.',
            'error',
        )
    else:
        flash('AI API key(s) saved successfully!', 'success')

    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-google-maps', methods=['POST'])
def save_google_maps():
    """Save Google Maps API keys."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))
    browser_key = request.form.get('google_maps_browser_api_key', '').strip()
    server_key = request.form.get('google_maps_server_api_key', '').strip()
    saved_any = False
    failed_any = False

    if browser_key:
        if set_setting('google_maps_browser_api_key', browser_key):
            saved_any = True
        else:
            failed_any = True
            logger.error("Failed to save Google Maps browser API key")
    if server_key:
        if set_setting('google_maps_server_api_key', server_key):
            saved_any = True
            # Keep legacy single-key setting in sync for older code paths/deployments.
            if not set_setting('google_maps_api_key', server_key):
                logger.warning("Failed to sync legacy google_maps_api_key")
        else:
            failed_any = True
            logger.error("Failed to save Google Maps server API key")

    if not browser_key and not server_key:
        flash('No Google Maps keys provided.', 'warning')
    elif failed_any and not saved_any:
        flash('Could not save Google Maps keys (database error). Check server logs and DATABASE_URL.', 'error')
    elif failed_any:
        flash('Some Google Maps keys were saved, but one key failed to save.', 'warning')
    else:
        flash('Google Maps API keys saved.', 'success')
    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-calendar', methods=['POST'])
def save_calendar():
    """Handle credentials.json upload for Google Vision and schedule API key."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    # Handle credentials.json file upload (used by Google Vision for deposit verification)
    credentials_upload_ok = False
    if 'credentials_file' in request.files:
        file = request.files['credentials_file']
        if file and file.filename and file.filename.endswith('.json'):
            filepath = CREDENTIALS_JSON_PATH
            try:
                content = file.read()
                json.loads(content)  # Validate JSON
                with open(filepath, 'wb') as f:
                    f.write(content)
                credentials_upload_ok = True
                logger.info("Vision credentials file uploaded")
            except json.JSONDecodeError:
                flash('Invalid JSON file', 'error')
            except Exception as e:
                flash(f'Error uploading file: {str(e)}', 'error')
                logger.error("Error uploading credentials: %s", e)
        elif file and file.filename:
            flash('File must be a .json file', 'warning')

    # Save schedule_api_key if provided (blank = keep existing)
    new_schedule_api_key = (request.form.get('schedule_api_key') or '').strip()
    if new_schedule_api_key:
        set_setting('schedule_api_key', new_schedule_api_key)
        logger.info("Schedule API key updated")
        flash('Schedule API key saved.', 'success')

    if credentials_upload_ok:
        flash('Google Vision credentials.json uploaded successfully.', 'success')

    return redirect(url_for('config.config_page'))


def _clamp_payid_base_hours(field_name: str, default: float) -> float:
    raw_value = request.form.get(field_name, default)
    value = float(raw_value)
    return max(0.5, min(12, value))


def _get_payid_form_data() -> dict[str, object]:
    return {
        'payid': request.form.get('payid', '').strip(),
        'account_name': request.form.get('account_name', '').strip(),
        'require_deposits': 'require_deposits' in request.form,
        'require_incall_deposits': 'require_incall_deposits' in request.form,
        'ugly_mugs_sync_enabled_present': request.form.get('ugly_mugs_sync_enabled_present', '').strip() == '1',
        'ugly_mugs_sync_enabled': 'ugly_mugs_sync_enabled' in request.form,
        'deposit_verification_vision': 'deposit_verification_vision' in request.form,
        'profanity_deposit_enabled': 'profanity_deposit_enabled' in request.form,
        'blocked_words_block_enabled': 'blocked_words_block_enabled' in request.form,
        'deposit_incall_scale': 'deposit_incall_scale_duration' in request.form,
        'deposit_outcall_scale': 'deposit_outcall_scale_duration' in request.form,
        'deposit_group_scale': 'deposit_group_scale_duration' in request.form,
        'deposit_dinner_date_outcall_scale': 'deposit_dinner_date_outcall_scale_duration' in request.form,
        'deposit_extended_experience_outcall_scale': 'deposit_extended_experience_outcall_scale_duration' in request.form,
    }


def _get_payid_pricing_values(pricing_defaults: dict[str, object]) -> dict[str, object]:
    return {
        'deposit_incall': int(str(request.form.get('deposit_incall', pricing_defaults.get('deposit_incall', 50)))),
        'deposit_outcall': int(str(request.form.get('deposit_outcall', pricing_defaults.get('deposit_outcall', 100)))),
        'deposit_group': int(str(request.form.get('deposit_group', pricing_defaults.get('deposit_mff_pair', 200)))),
        'deposit_dinner_date_outcall': int(
            str(request.form.get('deposit_dinner_date_outcall', pricing_defaults.get('deposit_dinner_date_outcall', 100)))
        ),
        'deposit_extended_experience_outcall': int(
            str(
                request.form.get(
                    'deposit_extended_experience_outcall',
                    pricing_defaults.get('deposit_extended_experience_outcall', 200),
                )
            )
        ),
        'deposit_incall_base_hours': _clamp_payid_base_hours('deposit_incall_base_hours', 1.0),
        'deposit_outcall_base_hours': _clamp_payid_base_hours('deposit_outcall_base_hours', 1.0),
        'deposit_group_base_hours': _clamp_payid_base_hours('deposit_group_base_hours', 2.0),
        'deposit_dinner_date_outcall_base_hours': _clamp_payid_base_hours('deposit_dinner_date_outcall_base_hours', 2.0),
        'deposit_extended_experience_outcall_base_hours': _clamp_payid_base_hours(
            'deposit_extended_experience_outcall_base_hours',
            2.0,
        ),
    }


def _save_payid_basic_settings(failed: list[str], form_data: dict[str, object]) -> None:
    operations = (
        ('PayID', 'payid', form_data['payid'], True),
        ('account name', 'account_name', form_data['account_name'], True),
        ('require deposits toggle', 'require_deposits', 'true' if form_data['require_deposits'] else 'false', True),
        (
            'require incall deposits toggle',
            'require_incall_deposits',
            'true' if form_data['require_incall_deposits'] else 'false',
            True,
        ),
        (
            'ugly mugs sync toggle',
            'ugly_mugs_sync_enabled',
            'true' if form_data['ugly_mugs_sync_enabled'] else 'false',
            bool(form_data['ugly_mugs_sync_enabled_present']),
        ),
    )
    for label, setting_key, setting_value, should_save in operations:
        _append_failed_setting(failed, label, setting_key, setting_value, should_save)


def _save_payid_pricing_settings(
    failed: list[str],
    form_data: dict[str, object],
    pricing_values: dict[str, object],
) -> None:
    operations = (
        ('incall deposit', 'deposit_incall', str(pricing_values['deposit_incall']), True),
        ('outcall deposit', 'deposit_outcall', str(pricing_values['deposit_outcall']), True),
        ('group deposit', 'deposit_group', str(pricing_values['deposit_group']), True),
        ('deposit_mff_pair (legacy sync)', 'deposit_mff_pair', str(pricing_values['deposit_group']), True),
        (
            'incall duration scaling toggle',
            'deposit_incall_scale_duration',
            'true' if form_data['deposit_incall_scale'] else 'false',
            True,
        ),
        ('incall base hours', 'deposit_incall_base_hours', str(pricing_values['deposit_incall_base_hours']), True),
        (
            'outcall duration scaling toggle',
            'deposit_outcall_scale_duration',
            'true' if form_data['deposit_outcall_scale'] else 'false',
            True,
        ),
        ('outcall base hours', 'deposit_outcall_base_hours', str(pricing_values['deposit_outcall_base_hours']), True),
        (
            'group duration scaling toggle',
            'deposit_group_scale_duration',
            'true' if form_data['deposit_group_scale'] else 'false',
            True,
        ),
        ('group base hours', 'deposit_group_base_hours', str(pricing_values['deposit_group_base_hours']), True),
        (
            'dinner date duration scaling toggle',
            'deposit_dinner_date_outcall_scale_duration',
            'true' if form_data['deposit_dinner_date_outcall_scale'] else 'false',
            True,
        ),
        (
            'dinner date base hours',
            'deposit_dinner_date_outcall_base_hours',
            str(pricing_values['deposit_dinner_date_outcall_base_hours']),
            True,
        ),
        (
            'extended experience duration scaling toggle',
            'deposit_extended_experience_outcall_scale_duration',
            'true' if form_data['deposit_extended_experience_outcall_scale'] else 'false',
            True,
        ),
        (
            'extended experience base hours',
            'deposit_extended_experience_outcall_base_hours',
            str(pricing_values['deposit_extended_experience_outcall_base_hours']),
            True,
        ),
    )
    for label, setting_key, setting_value, should_save in operations:
        _append_failed_setting(failed, label, setting_key, setting_value, should_save)


def _save_payid_pricing_config(failed: list[str], load_pricing, pricing_values: dict[str, object]) -> None:
    pricing_config = load_pricing()
    pricing_config['deposit_incall'] = pricing_values['deposit_incall']
    pricing_config['deposit_outcall'] = pricing_values['deposit_outcall']
    pricing_config['deposit_mff_pair'] = pricing_values['deposit_group']
    pricing_config['deposit_dinner_date_outcall'] = pricing_values['deposit_dinner_date_outcall']
    pricing_config['deposit_extended_experience_outcall'] = pricing_values['deposit_extended_experience_outcall']
    _append_failed_setting(failed, 'pricing config JSON', 'pricing_config', json.dumps(pricing_config))


def _save_payid_toggle_settings(failed: list[str], form_data: dict[str, object]) -> None:
    operations = (
        (
            'vision verification toggle',
            'deposit_verification_vision',
            'true' if form_data['deposit_verification_vision'] else 'false',
            True,
        ),
        (
            'profanity deposit toggle',
            'profanity_deposit_enabled',
            'true' if form_data['profanity_deposit_enabled'] else 'false',
            True,
        ),
        (
            'blocked words toggle',
            'blocked_words_block_enabled',
            'true' if form_data['blocked_words_block_enabled'] else 'false',
            True,
        ),
    )
    for label, setting_key, setting_value, should_save in operations:
        _append_failed_setting(failed, label, setting_key, setting_value, should_save)


@config_bp.route('/config/save-payid', methods=['POST'])
def save_payid():
    """Save PayID and deposit settings."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))

    form_data = _get_payid_form_data()
    try:
        from core.rates_from_config import _load_pricing, get_default_pricing

        pricing_values = _get_payid_pricing_values(get_default_pricing())
        failed: list[str] = []
        _save_payid_basic_settings(failed, form_data)
        _save_payid_pricing_settings(failed, form_data, pricing_values)
        _save_payid_pricing_config(failed, _load_pricing, pricing_values)
        _save_payid_toggle_settings(failed, form_data)
        if failed:
            flash(
                'Some payment settings could not be saved: ' + ', '.join(failed) + '. '
                'Check server logs and DATABASE_URL.',
                'error',
            )
            logger.error("save_payid: failed to save fields: %s", failed)
        else:
            flash('Payment settings saved successfully!', 'success')
            logger.info("Payment and deposit settings updated")
    except ValueError:
        flash('Invalid payment settings — please check deposit amounts and base hours.', 'error')

    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/debug/booking-mode-check', methods=['GET'])
def debug_booking_mode_check():
    """TEMP DIAGNOSTIC — remove after verifying incall_only is wired up."""
    try:
        val = get_setting('booking_mode')
        return jsonify({"booking_mode": val, "raw_type": type(val).__name__})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@config_bp.route('/config/save-booking-mode', methods=['POST'])
@require_auth
def save_booking_mode():
    """Save booking mode preference."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))

    booking_mode = (request.form.get('booking_mode') or '').strip()
    if booking_mode not in {'incall_outcall', 'incall_only'}:
        flash('Invalid booking mode selected.', 'error')
        return redirect(url_for('admin.admin_dashboard'))

    if set_setting('booking_mode', booking_mode):
        flash('Booking mode saved successfully!', 'success')
    else:
        flash('Could not save booking mode. Check server logs and DATABASE_URL.', 'error')

    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-hours', methods=['POST'])
def save_hours():
    """Save available hours and days settings."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))

    available_hours = request.form.get('available_hours', '').strip()
    available_days = request.form.get('available_days', '').strip()

    _failed = []
    if available_hours:
        if not set_setting('available_hours', available_hours):
            _failed.append('available hours')
    if available_days:
        if not set_setting('available_days', available_days):
            _failed.append('available days')

    if not available_hours and not available_days:
        flash('Please provide at least hours or days.', 'error')
    elif _failed:
        flash(
            'Could not save to database: ' + ', '.join(_failed) + '. Check server logs and DATABASE_URL.',
            'error',
        )
        logger.error("save_hours: failed to save fields: %s", _failed)
    else:
        flash('Availability settings updated successfully!', 'success')
        logger.info("Available hours: %s, days: %s", available_hours, available_days)

    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-escort-name', methods=['POST'])
def save_escort_name():
    """Save escort/business display name (used in all client messages and admin)."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))
    name = request.form.get('escort_name', '').strip()
    if not name:
        flash('Escort name cannot be empty', 'error')
        return redirect(url_for('admin.admin_dashboard'))
    if set_setting('escort_name', name):
        flash(f'Escort name updated to "{name}". All client-facing messages will use this name.', 'success')
        logger.info("Escort name updated to: %s", name)
    else:
        flash('Could not save escort name (database error). Check server logs and DATABASE_URL.', 'error')
        logger.error("save_escort_name: set_setting failed for escort_name")
    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-outcall-verification', methods=['POST'])
def save_outcall_verification():
    """Save outcall verification strict/lenient mode setting."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))

    strict_mode = 'outcall_verification_strict' in request.form
    mode = 'strict' if strict_mode else 'lenient'
    if set_setting('outcall_verification_strict', 'true' if strict_mode else 'false'):
        flash(f'Outcall verification mode set to {mode}.', 'success')
        logger.info("Outcall verification mode updated to: %s", mode)
    else:
        flash('Could not save outcall verification mode (database error). Check server logs and DATABASE_URL.', 'error')
        logger.error("save_outcall_verification: set_setting failed")

    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-escort-sms', methods=['POST'])
def save_escort_sms():
    """Save master escort SMS toggle and all per-category toggles."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))
    _failed = []
    # Master toggle
    enabled = 'escort_sms_enabled' in request.form
    if not set_setting('escort_sms_enabled', 'true' if enabled else 'false'):
        _failed.append('master SMS toggle')
    # Per-category toggles
    for cat_key, cat_label in ESCORT_SMS_CATEGORIES:
        val = f'escort_sms_{cat_key}' in request.form
        if not set_setting(f'escort_sms_{cat_key}', 'true' if val else 'false'):
            _failed.append(cat_label)
    # Client feedback flow (sync keys used by send_escort_sms category=client_rating / feedback_replies)
    client_feedback = 'client_feedback_enabled' in request.form
    if not set_setting('client_feedback_enabled', 'true' if client_feedback else 'false'):
        _failed.append('client feedback toggle')
    if not set_setting('escort_sms_client_rating', 'true' if client_feedback else 'false'):
        _failed.append('escort_sms_client_rating sync')
    if not set_setting('escort_sms_feedback_replies', 'true' if client_feedback else 'false'):
        _failed.append('escort_sms_feedback_replies sync')
    if not set_setting('escort_sms_feedback_requests', 'true' if client_feedback else 'false'):
        _failed.append('escort_sms_feedback_requests sync')
    incall_1h = 'incall_1h_reminder_enabled' in request.form
    if not set_setting('incall_1h_reminder_enabled', 'true' if incall_1h else 'false'):
        _failed.append('incall 1h client reminder')
    incall_fwd = 'incall_reminder_forward_replies' in request.form
    if not set_setting('incall_reminder_forward_replies', 'true' if incall_fwd else 'false'):
        _failed.append('incall forward replies to escort')
    if _failed:
        flash(
            'Some escort SMS settings could not be saved: ' + ', '.join(_failed) + '. '
            'Check server logs and DATABASE_URL.',
            'error',
        )
        logger.error("save_escort_sms: failed to save fields: %s", _failed)
    else:
        flash('Escort SMS settings saved.', 'success')
        logger.info("Escort SMS settings updated")
    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-base-url', methods=['POST'])
def save_base_url():
    """Save Site URL (base URL for deposit upload links and feedback links)."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    base_url = request.form.get('base_url', '').strip().rstrip('/')
    if set_setting('base_url', base_url):
        flash('Site URL saved. Deposit and feedback links will use this URL.', 'success')
        logger.info("Base URL updated: %s", base_url or "(empty)")
    else:
        flash('Could not save Site URL (database error). Check server logs and DATABASE_URL.', 'error')
        logger.error("Failed to save Site URL")

    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-redis-url', methods=['POST'])
def save_redis_url():
    """Save Redis URL for distributed rate limiting."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    redis_url = request.form.get('redis_url', '').strip()
    if redis_url and not (redis_url.startswith('redis://') or redis_url.startswith('rediss://')):
        flash('Invalid Redis URL — must start with redis:// or rediss:// (use rediss:// for Upstash TLS).', 'error')
        return redirect(url_for('config.config_page'))

    if set_setting('redis_url', redis_url):
        if redis_url:
            flash('Redis URL saved. Reload the app to activate persistent rate limiting.', 'success')
        else:
            flash('Redis URL cleared. Rate limiting will use in-memory fallback.', 'success')
        logger.info("Redis URL updated (length=%d)", len(redis_url))
    else:
        flash('Could not save Redis URL (database error).', 'error')

    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-startup-db-migrations', methods=['POST'])
def save_startup_db_migrations():
    """Toggle startup DB migrations for next app boot/reload."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    enabled = 'run_startup_db_migrations' in request.form
    if set_setting('run_startup_db_migrations', 'true' if enabled else 'false'):
        if enabled:
            flash(
                'Startup DB migrations enabled. Reload the web app once, then set this back to OFF.',
                'success',
            )
        else:
            flash('Startup DB migrations disabled.', 'success')
        logger.info("run_startup_db_migrations updated: %s", enabled)
    else:
        flash('Could not save startup migration setting (database error).', 'error')
        logger.error("Failed to save run_startup_db_migrations setting")
    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-flow-version-rollout', methods=['POST'])
def save_flow_version_rollout():
    """Save conversation flow routing controls (global force v1/v2 or rollout mode)."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    mode = (request.form.get('flow_version_default') or '').strip().lower()
    if mode not in ('rollout', 'v1', 'v2'):
        mode = 'rollout'

    raw_percent = (request.form.get('flow_version_v2_rollout_percent') or '').strip()
    try:
        percent = int(raw_percent) if raw_percent else 0
    except ValueError:
        percent = 0
    percent = max(0, min(100, percent))

    failed = []
    if not set_setting('flow_version_default', mode):
        failed.append('flow_version_default')
    if not set_setting('flow_version_v2_rollout_percent', str(percent)):
        failed.append('flow_version_v2_rollout_percent')

    if failed:
        flash(
            'Could not save flow rollout settings: ' + ', '.join(failed) + '.',
            'error',
        )
        logger.error("Failed to save flow rollout settings: %s", failed)
    else:
        flash(
            f'Flow rollout settings saved (mode={mode}, v2 rollout={percent}%).',
            'success',
        )
        logger.info(
            "flow rollout updated: mode=%s, v2_rollout_percent=%s",
            mode,
            percent,
        )
    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-ai-fallback-threshold', methods=['POST'])
def save_ai_fallback_threshold():
    """Save fallback confidence threshold (0.00 - 1.00)."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))

    def _sanitize_threshold(raw_value: str, default_value: float) -> str:
        try:
            parsed = float(raw_value) if raw_value else default_value
        except (TypeError, ValueError):
            parsed = default_value
        parsed = max(0.0, min(1.0, parsed))
        return f"{parsed:.2f}"

    global_threshold = _sanitize_threshold(
        (request.form.get('ai_fallback_confidence_threshold') or '').strip(),
        0.45,
    )
    per_step_fields = (
        "ai_fallback_confidence_threshold_qualification",
        "ai_fallback_confidence_threshold_availability",
        "ai_fallback_confidence_threshold_screening",
        "ai_fallback_confidence_threshold_deposit",
        "ai_fallback_confidence_threshold_confirmation",
        "ai_fallback_confidence_threshold_follow_up",
    )

    failed: list[str] = []
    if not set_setting('ai_fallback_confidence_threshold', global_threshold):
        failed.append('ai_fallback_confidence_threshold')
    for setting_key in per_step_fields:
        threshold_text = _sanitize_threshold(
            (request.form.get(setting_key) or '').strip(),
            float(global_threshold),
        )
        if not set_setting(setting_key, threshold_text):
            failed.append(setting_key)

    if failed:
        flash(
            'Could not save AI fallback confidence thresholds: ' + ', '.join(failed) + '.',
            'error',
        )
        logger.error("Failed to save AI fallback confidence thresholds: %s", failed)
    else:
        flash('AI fallback confidence thresholds saved.', 'success')
        logger.info("ai_fallback_confidence_thresholds updated (global=%s)", global_threshold)
    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-conversation-timeout', methods=['POST'])
def save_conversation_timeout():
    """Save conversation timeout (hours of inactivity before state reset)."""
    if not _is_admin_or_config_authenticated():
        return redirect(url_for('admin.admin_dashboard'))
    raw = request.form.get('conversation_timeout_hours', '').strip()
    try:
        hours = int(raw) if raw else 24
        hours = max(1, min(168, hours))  # 1 to 168 (1 week)
    except ValueError:
        hours = 24
    if set_setting('conversation_timeout_hours', str(hours)):
        flash(f'Conversation timeout set to {hours} hours.', 'success')
        logger.info("Conversation timeout hours updated: %s", hours)
    else:
        flash('Could not save conversation timeout (database error). Check server logs and DATABASE_URL.', 'error')
        logger.error("save_conversation_timeout: set_setting failed")
    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-escort-phone', methods=['POST'])
def save_escort_phone():
    """Save escort mobile number used for SMS notifications."""
    if not _is_admin_or_config_authenticated():
        flash('You must be logged in to save escort phone.', 'error')
        return redirect(url_for('admin.admin_dashboard'))
    phone = request.form.get('escort_phone_number', '').strip()
    if not phone:
        flash('Phone number cannot be empty.', 'warning')
        return redirect(url_for('admin.admin_dashboard'))
    try:
        if set_setting('escort_phone_number', phone):
            flash('Escort phone number saved.', 'success')
            logger.info("Escort phone number updated")
        else:
            flash(
                'Could not save escort phone (database error). Check DATABASE_URL and admin_settings.',
                'error',
            )
            logger.error("set_setting failed for escort_phone_number")
    except Exception as e:
        logger.exception("save_escort_phone failed: %s", e)
        flash(
            f'Could not save escort phone: {e!s}. Check server logs and DATABASE_URL.',
            'error',
        )
    return redirect(url_for('admin.admin_dashboard'))


@config_bp.route('/config/save-flask-secret', methods=['POST'])
def save_flask_secret():
    """Save Flask session secret key."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))
    key = request.form.get('flask_secret_key', '').strip()
    if key:
        if set_setting('flask_secret_key', key):
            flash(
                'Flask secret key saved. Reload the web app in PythonAnywhere for it to take effect.',
                'success',
            )
            logger.info("Flask secret key updated")
        else:
            flash(
                'Could not save Flask secret key (database error). Check server logs and DATABASE_URL.',
                'error',
            )
    else:
        flash('Secret key cannot be empty.', 'warning')
    return redirect(url_for('config.config_page'))


@config_bp.route('/config/save-opencage', methods=['POST'])
def save_opencage():
    """Save OpenCage geocoding API key (empty string clears the stored key)."""
    if not _is_config_authenticated():
        return redirect(url_for('config.config_page'))
    key = request.form.get('opencage_api_key', '').strip()
    if set_setting('opencage_api_key', key):
        flash('OpenCage API key saved.' if key else 'OpenCage API key cleared.', 'success')
        logger.info("OpenCage API key updated (set=%s)", bool(key))
    else:
        flash('Could not save OpenCage API key (database error). Check server logs and DATABASE_URL.', 'error')
        logger.error("Failed to save OpenCage API key")
    return redirect(url_for('config.config_page'))

