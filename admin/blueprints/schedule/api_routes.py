"""Schedule JSON APIs and logout."""



import hashlib
import json as _json
import threading
import time
from datetime import datetime, timedelta
from types import SimpleNamespace

import pytz
from flask import jsonify, make_response, request, session
from werkzeug.datastructures import MultiDict

import config
from admin.auth import verify_password
from core.settings_manager import set_setting
from services.database_service import get_shared_db
from services.push_notification_service import (
    _list_active_tokens,
    _send_token,
    register_push_device_token,
    send_deposit_paid_push_for_booking_id,
    send_new_booking_push_for_booking_id,
)
from services.safety_screening_service import lookup_flagged_number
from utils.row_utils import row_get

from .blueprint import schedule_bp
from booking.mmf_exploration import schedule_should_show_mmf_preferences


_MOBILE_SYNC_CACHE_TTL_SECONDS = 30
_MOBILE_SYNC_CACHE_MAX_ENTRIES = 20
_MOBILE_SYNC_CACHE: dict[tuple[str, str], tuple[float, dict]] = {}
_MOBILE_SYNC_CACHE_LOCK = threading.Lock()


def _mobile_sync_cache_get(key: tuple[str, str]) -> dict | None:
    now = time.monotonic()
    with _MOBILE_SYNC_CACHE_LOCK:
        entry = _MOBILE_SYNC_CACHE.get(key)
        if entry and (now - entry[0]) < _MOBILE_SYNC_CACHE_TTL_SECONDS:
            return entry[1]
        if entry:
            _MOBILE_SYNC_CACHE.pop(key, None)
    return None


def _mobile_sync_cache_set(key: tuple[str, str], payload: dict) -> None:
    with _MOBILE_SYNC_CACHE_LOCK:
        _MOBILE_SYNC_CACHE[key] = (time.monotonic(), payload)
        if len(_MOBILE_SYNC_CACHE) > _MOBILE_SYNC_CACHE_MAX_ENTRIES:
            oldest = min(_MOBILE_SYNC_CACHE.items(), key=lambda kv: kv[1][0])[0]
            _MOBILE_SYNC_CACHE.pop(oldest, None)


def _mobile_sync_cache_invalidate() -> None:
    with _MOBILE_SYNC_CACHE_LOCK:
        _MOBILE_SYNC_CACHE.clear()


def _jsonify_mobile_sync_with_etag(payload):
    """Return JSON response with ETag; 304 if the APK already has this version."""
    raw = _json.dumps(payload, sort_keys=True, separators=(",", ":"))
    etag = '"' + hashlib.md5(raw.encode()).hexdigest() + '"'
    if request.headers.get("If-None-Match") == etag:
        return make_response("", 304)
    resp = make_response(raw, 200)
    resp.headers["Content-Type"] = "application/json"
    resp.headers["ETag"] = etag
    return resp



@schedule_bp.route('/schedule/api/messages', methods=['GET'])
def get_messages_api():
    """Get SMS message history for a phone number (Bearer auth)."""
    if not _is_schedule_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    phone = request.args.get("phone", "").strip()
    if not phone:
        return jsonify({"error": "phone parameter required"}), 400

    phone_digits = "".join(ch for ch in phone if ch.isdigit())
    if not phone_digits:
        return jsonify({"error": "phone parameter required"}), 400

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503
        suffix = phone_digits[-9:]
        rows = db.execute_query(
            """
            SELECT id, direction, message_body, created_at, phone_number
            FROM message_history
            WHERE regexp_replace(COALESCE(phone_number, ''), '\\D', '', 'g') LIKE %s
            ORDER BY created_at DESC
            LIMIT 100
            """,
            (f"%{suffix}",),
            fetch=True,
        ) or []

        messages = []
        for row in rows:
            if isinstance(row, dict):
                timestamp = row.get("created_at")
                messages.append({
                    "id": row.get("id"),
                    "direction": row.get("direction") or "",
                    "body": row.get("message_body") or "",
                    "timestamp": timestamp.isoformat() if timestamp is not None and hasattr(timestamp, "isoformat") else (str(timestamp) if timestamp else None),
                    "phone_number": row.get("phone_number") or "",
                })
            else:
                timestamp = row_get(row, 3, None)
                messages.append({
                    "id": row_get(row, 0, None),
                    "direction": row_get(row, 1, "") or "",
                    "body": row_get(row, 2, "") or "",
                    "timestamp": timestamp.isoformat() if timestamp is not None and hasattr(timestamp, "isoformat") else (str(timestamp) if timestamp else None),
                    "phone_number": row_get(row, 4, "") or "",
                })

        return jsonify({"messages": messages})
    except Exception as e:
        logger.error("Error fetching messages for %s: %s", phone, e, exc_info=True)
        return jsonify({"error": "Internal server error"}), 500

_STATUS_TO_COLOR_ID = {
    "confirmed": "10", "reschedule-confirmed": "10",
    "pending": "8", "pending-deposit": "8",
    "reserved": "7", "travel": "3", "pending-travel": "1",
    "admin": "5", "social": "11",
}


def _effective_booking_status(status: str, deposit_status: str) -> str:
    normalized_status = (status or "").strip().lower()
    normalized_dep = (deposit_status or "").strip().lower()
    if normalized_dep == "pending" and normalized_status in {"", "reserved", "pending", "pending-deposit"}:
        return "pending-deposit"
    return normalized_status or "reserved"


def _api_parse_mobile_sync_range():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    tz = _get_local_timezone()
    try:
        if start_date:
            start_dt = tz.localize(datetime.strptime(start_date, "%Y-%m-%d"))
        else:
            start_dt = tz.localize((datetime.now() - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0))
        if end_date:
            end_dt = tz.localize(datetime.strptime(end_date, "%Y-%m-%d")).replace(hour=23, minute=59, second=59)
        else:
            end_dt = tz.localize((datetime.now() + timedelta(days=90)).replace(hour=23, minute=59, second=59))
    except Exception:
        return None, None, (jsonify({"error": "Invalid date format"}), 400)
    return start_dt, end_dt, None



def _api_fetch_mobile_sync_rows(db, start_dt, end_dt):
    return db.execute_query(
        """
        SELECT id, start_time, end_time, client_name, phone, duration, type,
               experience, preferences, deposit_status, deposit_amount,
               deposit_reference, status, special_requests, organise_other_escort,
               notes, price_total, remaining_amount, outcall_address
        FROM bookings
        WHERE start_time >= %s AND start_time <= %s
        ORDER BY start_time ASC
        LIMIT 2500
        """,
        (start_dt.isoformat(), end_dt.isoformat()),
        fetch=True,
    ) or []



def _api_get_mobile_sync_times(row):
    start = row_get(row, "start_time")
    end = row_get(row, "end_time")
    return {
        "start": start.isoformat() if hasattr(start, "isoformat") else str(start or ""),
        "end": end.isoformat() if hasattr(end, "isoformat") else str(end or ""),
        "duration": str(row_get(row, "duration") or ""),
    }



def _api_build_mobile_sync_summary_fields(row, booking_id, client_name, experience, loc_type, status):
    return {
        "event_id": booking_id,
        "summary": f"{client_name} — {experience or loc_type}".rstrip(" —").strip(),
        "description": "",
        "color_id": _STATUS_TO_COLOR_ID.get(status, "8"),
        "phone_number": str(row_get(row, "phone") or ""),
        "client_name": client_name,
        "experience": experience,
        "organise_other_escort": "yes" if row_get(row, "organise_other_escort") else "no",
        "safety_screening_status": "",
        "location_type": loc_type,
        "address": str(row_get(row, "outcall_address") or ""),
        "origin_address": "",
        "destination_address": "",
        "status_class": status,
        "status_label": status.replace("-", " ").title(),
        "original_datetime_display": "",
        "notes": str(row_get(row, "notes") or ""),
    }



def _api_build_mobile_sync_financial_fields(row, dep_status, dep_amount, price_total, remaining):
    return {
        "price": str(float(price_total)) if price_total is not None else "",
        "deposit_paid": str(dep_amount) if dep_status == "paid" else "0",
        "deposit_due": str(dep_amount) if dep_status == "pending" else "",
        "deposit_reference": str(row_get(row, "deposit_reference") or ""),
        "remaining_balance": str(float(remaining)) if remaining is not None else "",
        "special_requests": str(row_get(row, "special_requests") or "").strip(),
    }



def _api_build_mobile_sync_preference_fields(prefs):
    return {
        "preferences": list(prefs) if isinstance(prefs, (list, tuple)) else [],
        "show_mmf_preferences": False,
    }



def _api_serialize_mobile_sync_booking(row):
    booking_id = str(row_get(row, "id") or "")
    raw_status = str(row_get(row, "status") or "reserved")
    dep_status = str(row_get(row, "deposit_status") or "not_required")
    status = _effective_booking_status(raw_status, dep_status)
    dep_amount = float(row_get(row, "deposit_amount") or 0)
    prefs = row_get(row, "preferences") or []
    client_name = str(row_get(row, "client_name") or "Client")
    experience = str(row_get(row, "experience") or "").replace("_", " ")
    loc_type = str(row_get(row, "type") or "")
    price_total = row_get(row, "price_total")
    remaining = row_get(row, "remaining_amount")
    booking = _api_get_mobile_sync_times(row)
    booking.update(_api_build_mobile_sync_summary_fields(row, booking_id, client_name, experience, loc_type, status))
    booking.update(_api_build_mobile_sync_financial_fields(row, dep_status, dep_amount, price_total, remaining))
    booking.update(_api_build_mobile_sync_preference_fields(prefs))
    return booking



def _api_build_mobile_sync_bookings(rows):
    bookings = []
    for row in rows:
        try:
            bookings.append(_api_serialize_mobile_sync_booking(row))
        except Exception as row_err:
            logger.warning("mobile-sync row serialise error: %s", row_err)
            continue
    return bookings



def _api_get_mobile_sync_timezone():
    try:
        return _get_local_timezone().zone
    except Exception:
        return "Australia/Adelaide"



@schedule_bp.route('/schedule/api/mobile-sync', methods=['GET'])
def schedule_api_mobile_sync():
    """Mobile APK sync — reads from the bookings DB table (no longer uses Google Calendar)."""
    if not _is_schedule_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    start_dt, end_dt, error_response = _api_parse_mobile_sync_range()
    if error_response:
        return error_response
    if start_dt is None or end_dt is None:
        return jsonify({"error": "Invalid date format"}), 400

    cache_key = (start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"))
    cached_payload = _mobile_sync_cache_get(cache_key)
    if cached_payload is not None:
        return _jsonify_mobile_sync_with_etag(cached_payload)

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503
        rows = _api_fetch_mobile_sync_rows(db, start_dt, end_dt)
    except Exception as e:
        logger.error("mobile-sync DB query failed: %s", e)
        return jsonify({"error": "Database error"}), 500

    payload = {
        "bookings": _api_build_mobile_sync_bookings(rows),
        "timezone": _api_get_mobile_sync_timezone(),
    }
    _mobile_sync_cache_set(cache_key, payload)
    return _jsonify_mobile_sync_with_etag(payload)


def _api_first_present(*values):
    for value in values:
        if value:
            return value
    return ""



def _api_get_push_token_value(data, primary_key, secondary_key):
    return str(_api_first_present(data.get(primary_key), data.get(secondary_key), "")).strip()



def _api_get_push_token_metadata(data):
    return {
        "user_agent": request.headers.get("User-Agent") or "",
        "device_name": str(_api_first_present(data.get("deviceName"), data.get("device_name"), ""))[:120],
        "app_version": str(_api_first_present(data.get("appVersion"), data.get("app_version"), ""))[:40],
        "build_number": str(_api_first_present(data.get("buildNumber"), data.get("build_number"), ""))[:40],
        "incallChannelId": str(_api_first_present(data.get("incallChannelId"), ""))[:80],
        "outcallChannelId": str(_api_first_present(data.get("outcallChannelId"), ""))[:80],
        "channelId": str(_api_first_present(data.get("channelId"), ""))[:80],
    }



def _api_get_push_token_request_data():
    data = request.get_json(silent=True) or {}
    return {
        "expo_push_token": _api_get_push_token_value(data, "expoPushToken", "expo_push_token"),
        "fcm_token": _api_get_push_token_value(data, "fcmToken", "fcm_token"),
        "platform": str(_api_first_present(data.get("platform"), "android")).strip().lower() or "android",
        "provider": str(_api_first_present(data.get("provider"), "fcm")).strip().lower() or "fcm",
        "metadata": _api_get_push_token_metadata(data),
    }



def _api_validate_push_token_request(push_data):
    if not push_data["expo_push_token"] and not push_data["fcm_token"]:
        return jsonify({"error": "expoPushToken or fcmToken is required"}), 400
    return None



@schedule_bp.route("/schedule/api/push-token", methods=["POST", "PUT"])
@schedule_bp.route("/schedule/api/mobile/push-token", methods=["POST", "PUT"])
@schedule_bp.route("/schedule/api/device-token", methods=["POST", "PUT"])
@schedule_bp.route("/api/push-token", methods=["POST", "PUT"])
@schedule_bp.route("/api/device-token", methods=["POST", "PUT"])
def register_push_token_api():
    """Register Expo/FCM device tokens for server-side push notifications."""
    if not _is_schedule_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    push_data = _api_get_push_token_request_data()
    error_response = _api_validate_push_token_request(push_data)
    if error_response:
        return error_response

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503

        registered_count = register_push_device_token(
            db,
            platform=push_data["platform"],
            provider=push_data["provider"],
            expo_push_token=push_data["expo_push_token"],
            fcm_token=push_data["fcm_token"],
            metadata=push_data["metadata"],
        )
        return jsonify({"success": True, "registered": registered_count}), 200
    except Exception as e:
        logger.error("Push token registration failed: %s", e, exc_info=True)
        return jsonify({"error": "Failed to register push token"}), 500


@schedule_bp.route("/schedule/api/push-debug", methods=["GET", "POST"])
def push_debug_api():
    """Debug endpoint: list registered push tokens and optionally send a test push."""
    if not _is_schedule_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    db = get_shared_db(config.DATABASE_URL)
    if db is None:
        return jsonify({"error": "Database unavailable"}), 503

    from services.push_notification_service import _ensure_push_schema
    _ensure_push_schema(db)

    rows = db.execute_query(
        """
        SELECT token, token_type, platform, active, last_seen_at, last_error, updated_at
        FROM push_device_tokens
        ORDER BY updated_at DESC
        """,
        fetch=True,
    ) or []

    tokens_info = []
    for r in rows:
        tok = str(row_get(r, "token") or "")
        masked = tok[:12] + "…" + tok[-6:] if len(tok) > 20 else tok
        tokens_info.append({
            "token_preview": masked,
            "token_type": row_get(r, "token_type"),
            "platform": row_get(r, "platform"),
            "active": row_get(r, "active"),
            "last_seen_at": str(row_get(r, "last_seen_at") or ""),
            "last_error": row_get(r, "last_error"),
            "updated_at": str(row_get(r, "updated_at") or ""),
        })

    result = {"token_count": len(tokens_info), "tokens": tokens_info}

    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        if data.get("action") == "test":
            active_rows = _list_active_tokens(db)
            if not active_rows:
                result["test_push"] = {"sent": 0, "error": "No active tokens registered"}
            else:
                sent = 0
                errors = []
                for r in active_rows:
                    token = str(row_get(r, "token") or "").strip()
                    token_type = str(row_get(r, "token_type") or "").strip().lower()
                    if not token or token_type not in {"expo", "fcm"}:
                        continue
                    ok, err, _ = _send_token(
                        token=token,
                        token_type=token_type,
                        title="Test Push",
                        body="Push notifications are working!",
                        data={"kind": "test"},
                    )
                    if ok:
                        sent += 1
                    else:
                        errors.append(err)
                result["test_push"] = {"sent": sent, "errors": errors}

    return jsonify(result), 200


from .helpers import (
    _get_local_timezone,
    _is_schedule_authenticated,
)
from .log import logger
from .page_routes import _delete_travel_time_blocks, _handle_cancellation, _handle_reschedule


def _is_doubles_experience(experience: str | None) -> bool:
    value = (experience or "").strip().lower()
    return value in {"doubles mff", "doubles mmf", "doubles_mff", "Doubles MMF"}


def _normalize_organise_other_escort(value) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in {"yes", "y", "true", "1"}:
        return "yes"
    if normalized in {"no", "n", "false", "0"}:
        return "no"
    return None


def _resolve_safety_screening_status(phone_number: str | None) -> str | None:
    phone = (phone_number or "").strip()
    if not phone:
        return None
    try:
        lookup = lookup_flagged_number(phone)
        if lookup.get("matched"):
            return "flagged watchlist match"
    except Exception as e:
        logger.warning("Schedule safety screening lookup skipped for %s: %s", phone, e)
    return None


@schedule_bp.route("/schedule/api/update-availability", methods=["POST"])
def schedule_api_update_availability():
    """Update available hours/days from the schedule page. Accepts schedule or password auth."""
    # Allow either schedule session or password in form
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        password = request.form.get("password", "").strip()
        if not password or not verify_password(password):
            return jsonify({"success": False, "error": "Unauthorized. Please log in or enter the correct password."}), 401

    available_hours = (request.form.get("available_hours") or "").strip()
    if available_hours.lower() in ("no days selected", ", no days selected"):
        return jsonify({"success": False, "error": "Please select at least one day."}), 400
    # Blank time range = 24/7
    if not available_hours or available_hours.startswith("24/7"):
        available_hours = "24/7, 7 days a week"

    # Persists to admin_settings (setting_key=available_hours) — same table as /config/save-hours.
    if not set_setting("available_hours", available_hours):
        logger.error("set_setting failed for available_hours (database unreachable or error)")
        return jsonify({
            "success": False,
            "error": "Could not save to database. Check DATABASE_URL and server logs.",
        }), 500
    logger.info("Available hours updated from schedule page: %s", available_hours)
    return jsonify({
        "success": True,
        "message": "Available hours updated successfully",
        "available_hours": available_hours,
    })


def _api_parse_optional_float(value, empty_values=(None, "")):
    try:
        return float(value) if value not in empty_values else None
    except (ValueError, TypeError):
        return None



def _api_normalize_preferences(preferences):
    if isinstance(preferences, list):
        return [str(p).strip() for p in preferences if str(p).strip()]
    if preferences:
        return [str(preferences).strip()]
    return []



def _api_localize_event_window(date_str, time_str, duration_hours, tz):
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    time_obj = datetime.strptime(time_str, "%H:%M")
    start_dt = tz.localize(datetime(
        date_obj.year, date_obj.month, date_obj.day,
        time_obj.hour, time_obj.minute,
    ))
    return start_dt, start_dt + timedelta(hours=duration_hours)



def _api_fetch_schedule_bookings_rows(db, tz_name, selected_date):
    return db.execute_query(
        """
        SELECT id, start_time, end_time, client_name, phone, duration, type,
               experience, preferences, deposit_status, deposit_amount,
               deposit_reference, status, special_requests, organise_other_escort,
               notes, price_total, remaining_amount, outcall_address
        FROM bookings
        WHERE DATE(start_time AT TIME ZONE %s) = %s
        ORDER BY start_time ASC
        """,
        (tz_name, selected_date),
        fetch=True,
    ) or []



def _api_get_schedule_booking_times(row, tz):
    start = row_get(row, "start_time")
    end = row_get(row, "end_time")
    start_local = start.astimezone(tz) if getattr(start, "tzinfo", None) else tz.localize(start)
    end_local = end.astimezone(tz) if getattr(end, "tzinfo", None) else tz.localize(end)
    return {
        "start_time": start_local.strftime("%I:%M%p"),
        "end_time": end_local.strftime("%I:%M%p"),
        "duration": str(row_get(row, "duration") or ""),
    }



def _api_build_schedule_booking_detail_fields(row, booking_id, organise, pref_str, status):
    details_for_mmf = {
        "experience": str(row_get(row, "experience") or ""),
        "organise_other_escort": "yes" if organise else "no",
        "preferences": pref_str,
    }
    return {
        "event_id": booking_id,
        "phone_number": str(row_get(row, "phone") or ""),
        "client_name": str(row_get(row, "client_name") or "Client"),
        "experience": str(row_get(row, "experience") or ""),
        "organise_other_escort": "yes" if organise else "",
        "safety_screening_status": _resolve_safety_screening_status(str(row_get(row, "phone") or "")),
        "location_type": str(row_get(row, "type") or ""),
        "status_class": status,
        "special_requests": str(row_get(row, "special_requests") or "").strip(),
        "preferences": pref_str,
        "show_mmf_preferences": schedule_should_show_mmf_preferences(details_for_mmf),
    }



def _api_build_schedule_booking_financial_fields(row, dep_status, dep_amount, price_total, remaining):
    return {
        "price": str(float(price_total)) if price_total is not None else "",
        "deposit_paid": str(dep_amount) if dep_status == "paid" else "",
        "deposit_reference": str(row_get(row, "deposit_reference") or ""),
        "remaining_balance": str(float(remaining)) if remaining is not None else "",
    }



def _api_serialize_schedule_booking_row(row, tz):
    booking_id = str(row_get(row, "id") or "")
    raw_status = str(row_get(row, "status") or "reserved")
    dep_status = str(row_get(row, "deposit_status") or "not_required")
    status = _effective_booking_status(raw_status, dep_status)
    dep_amount = float(row_get(row, "deposit_amount") or 0)
    price_total = row_get(row, "price_total")
    remaining = row_get(row, "remaining_amount")
    prefs = row_get(row, "preferences") or []
    organise = row_get(row, "organise_other_escort")
    pref_str = ", ".join(prefs) if isinstance(prefs, (list, tuple)) else str(prefs or "")
    booking = _api_get_schedule_booking_times(row, tz)
    booking.update(_api_build_schedule_booking_detail_fields(row, booking_id, organise, pref_str, status))
    booking.update(_api_build_schedule_booking_financial_fields(row, dep_status, dep_amount, price_total, remaining))
    return booking



def _api_build_schedule_bookings(rows, tz):
    bookings = []
    for row in rows:
        try:
            bookings.append(_api_serialize_schedule_booking_row(row, tz))
        except Exception as row_err:
            logger.warning("Skipping malformed booking row %s: %s", row_get(row, "id", "?"), row_err)
    return bookings



@schedule_bp.route("/schedule/api/bookings", methods=["GET"])
def schedule_api_bookings():
    """API endpoint for fetching bookings for week view."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    selected_date = request.args.get("date")
    if not selected_date:
        return jsonify({"error": "Date parameter required"}), 400

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503
        tz = _get_local_timezone()
        rows = _api_fetch_schedule_bookings_rows(db, str(tz), selected_date)
        return jsonify({"bookings": _api_build_schedule_bookings(rows, tz)})
    except Exception as e:
        logger.error(f"Error in schedule API: {e}")
        return jsonify({"error": str(e)}), 500


def _api_get_create_event_data():
    data = request.get_json()
    return {
        "event_type": data.get("event_type"),
        "title": data.get("title"),
        "date_str": data.get("date"),
        "time_str": data.get("time"),
        "duration_hours": float(data.get("duration", 1)),
        "notes": data.get("notes", ""),
        "client_name": data.get("client_name", ""),
        "client_phone": data.get("client_phone", ""),
        "location_type": data.get("location_type", "incall"),
        "booking_status": data.get("booking_status", "peacock"),
        "deposit_amount": data.get("deposit_amount", "none"),
        "deposit_reference": str(data.get("deposit_reference") or "").strip(),
        "total_cost": data.get("total_cost"),
        "remaining_balance": data.get("remaining_balance"),
        "experience_type": data.get("experience_type", ""),
        "organise_other_escort": _normalize_organise_other_escort(data.get("organise_other_escort")),
        "preferences": data.get("preferences", []),
    }



def _api_validate_create_event_data(event_data):
    if not all([event_data["event_type"], event_data["date_str"], event_data["time_str"]]):
        return jsonify({"error": "Missing required fields"}), 400
    return None



def _api_resolve_create_event_title(event_data):
    title = event_data["title"]
    if event_data["event_type"] == "booking" and not title:
        return event_data["client_name"] or "Booking"
    return title



def _api_resolve_create_db_fields(event_type, location_type, booking_status):
    if event_type == "booking":
        if booking_status == "confirmed":
            return location_type or "incall", "confirmed"
        if booking_status == "pending_deposit":
            return location_type or "incall", "pending-deposit"
        return location_type or "incall", "reserved"
    return event_type, event_type



def _api_check_create_booking_conflicts(db, event_type, start_dt, end_dt):
    if event_type != "booking":
        return None
    blocking = ('confirmed', 'reschedule-confirmed', 'reserved', 'travel', 'admin', 'social')
    conflicts = db.execute_query(
        """
        SELECT client_name FROM bookings
        WHERE status = ANY(%s)
        AND start_time < %s AND end_time > %s
        """,
        (list(blocking), end_dt.isoformat(), start_dt.isoformat()),
        fetch=True,
    ) or []
    if conflicts:
        names = [str(row_get(r, "client_name") or "existing booking") for r in conflicts[:2]]
        return jsonify({"error": f"Time slot conflicts with: {', '.join(names)}. Please choose a different time."}), 409
    return None



def _api_resolve_create_dep_status(event_type, booking_status, dep_amount_val):
    if event_type != "booking":
        return "not_required"
    if booking_status == "confirmed" and dep_amount_val:
        return "paid"
    if booking_status == "pending_deposit":
        return "pending"
    return "not_required"



def _api_fill_create_price_total(event_data, price_total_val):
    if price_total_val is None and event_data["experience_type"] and event_data["duration_hours"]:
        try:
            from templates.confirmations import calculate_price
            c = int(calculate_price(
                int(event_data["duration_hours"] * 60),
                experience_type=event_data["experience_type"],
                incall_outcall=(event_data["location_type"] or "incall").lower(),
            ))
            if c > 0:
                return float(c)
        except Exception:
            pass
    return price_total_val



def _api_resolve_create_financials(event_data):
    dep_amount_val = _api_parse_optional_float(event_data["deposit_amount"], ("none", None, ""))
    price_total_val = _api_parse_optional_float(event_data["total_cost"])
    remaining_amount_val = _api_parse_optional_float(event_data["remaining_balance"])
    dep_status = _api_resolve_create_dep_status(
        event_data["event_type"],
        event_data["booking_status"],
        dep_amount_val,
    )
    if event_data["event_type"] == "booking":
        price_total_val = _api_fill_create_price_total(event_data, price_total_val)
        if price_total_val is not None and remaining_amount_val is None:
            remaining_amount_val = max(price_total_val - float(dep_amount_val or 0), 0)
    deposit_reference = event_data["deposit_reference"]
    if dep_status == "not_required":
        dep_amount_val = None
        deposit_reference = ""
    return {
        "dep_amount_val": dep_amount_val,
        "price_total_val": price_total_val,
        "remaining_amount_val": remaining_amount_val,
        "dep_status": dep_status,
        "deposit_reference": deposit_reference,
    }



def _api_ensure_legacy_booking_columns_nullable(db):
    for _ddl in (
        "ALTER TABLE bookings ALTER COLUMN date DROP NOT NULL",
        "ALTER TABLE bookings ALTER COLUMN time DROP NOT NULL",
        "ALTER TABLE bookings ALTER COLUMN phone_number DROP NOT NULL",
    ):
        try:
            db.execute_query(_ddl, (), fetch=False)
        except Exception:
            pass



def _api_insert_created_booking(db, event_data, db_type, db_status, start_dt, end_dt, financials):
    pref_list = _api_normalize_preferences(event_data["preferences"])
    _phone_val = (event_data["client_phone"] or '').strip()
    _local_dt = start_dt.astimezone(_get_local_timezone())
    result = db.execute_query(
        """
        INSERT INTO bookings (
            start_time, end_time, client_name, phone, phone_number, date, time, duration, type,
            experience, preferences, deposit_status, deposit_amount,
            deposit_reference, status, special_requests, organise_other_escort, notes,
            price_total, remaining_amount
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
        """,
        (
            start_dt.isoformat(), end_dt.isoformat(),
            event_data["client_name"] or None, _phone_val or None, _phone_val,
            _local_dt.strftime('%Y-%m-%d'), _local_dt.strftime('%H:%M'),
            int(event_data["duration_hours"] * 60), db_type,
            event_data["experience_type"] or None,
            pref_list,
            financials["dep_status"], financials["dep_amount_val"],
            financials["deposit_reference"] or None,
            db_status,
            event_data["notes"] or None,
            True if event_data["organise_other_escort"] == "yes" else (False if event_data["organise_other_escort"] == "no" else None),
            event_data["notes"] or None,
            financials["price_total_val"],
            financials["remaining_amount_val"],
        ),
        fetch=True,
    )
    return str(row_get(result[0], "id")) if result else None



def _api_send_create_booking_push(db, event_type, new_id):
    if event_type != "booking" or not new_id:
        return
    try:
        sent = send_new_booking_push_for_booking_id(db, new_id)
        if sent > 0:
            logger.info("Sent %s push notification(s) for new booking %s", sent, new_id)
    except Exception as push_err:
        logger.warning("New-booking push notification failed for %s: %s", new_id, push_err)



@schedule_bp.route("/schedule/api/create-event", methods=["POST"])
def create_event_api():
    """API endpoint for creating new bookings manually."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        event_data = _api_get_create_event_data()
        error_response = _api_validate_create_event_data(event_data)
        if error_response:
            return error_response

        title = _api_resolve_create_event_title(event_data)
        db_type, db_status = _api_resolve_create_db_fields(
            event_data["event_type"],
            event_data["location_type"],
            event_data["booking_status"],
        )
        tz = _get_local_timezone()
        start_dt, end_dt = _api_localize_event_window(
            event_data["date_str"],
            event_data["time_str"],
            event_data["duration_hours"],
            tz,
        )
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503

        conflict_response = _api_check_create_booking_conflicts(db, event_data["event_type"], start_dt, end_dt)
        if conflict_response:
            return conflict_response

        financials = _api_resolve_create_financials(event_data)
        _api_ensure_legacy_booking_columns_nullable(db)
        new_id = _api_insert_created_booking(db, event_data, db_type, db_status, start_dt, end_dt, financials)

        _mobile_sync_cache_invalidate()
        _api_send_create_booking_push(db, event_data["event_type"], new_id)
        return jsonify({
            "success": True,
            "event_id": new_id,
            "message": f"Event '{title}' created successfully"
        })

    except Exception as e:
        logger.exception("Error creating event")
        return jsonify({"error": str(e)}), 500


def _api_fetch_booking_for_update(db, event_id):
    rows = db.execute_query(
        """SELECT id, start_time, end_time, client_name, phone, duration, type,
                  experience, preferences, deposit_status, deposit_amount, status,
                  special_requests, organise_other_escort, notes, price_total, outcall_address
           FROM bookings WHERE id = %s""",
        (event_id,),
        fetch=True,
    ) or []
    return rows[0] if rows else None



def _api_prepare_update_timing(row, data, tz):
    existing_start = row_get(row, "start_time")
    existing_end = row_get(row, "end_time")
    existing_status = str(row_get(row, "status") or "reserved").strip().lower()
    duration_hours = data.get("duration")
    new_start = None
    new_end = None
    date_str = data.get("date")
    time_str = data.get("time")
    if date_str is not None and time_str is not None:
        hours = float(duration_hours) if duration_hours is not None else (
            (existing_end - existing_start).total_seconds() / 3600
            if existing_end and existing_start else 1.0
        )
        new_start, new_end = _api_localize_event_window(date_str, time_str, hours, tz)
    elif duration_hours is not None and existing_start:
        start_aware = existing_start.astimezone(tz) if getattr(existing_start, "tzinfo", None) else tz.localize(existing_start)
        new_start = start_aware
        new_end = new_start + timedelta(hours=float(duration_hours))
    return {
        "existing_status": existing_status,
        "duration_hours": duration_hours,
        "new_start": new_start,
        "new_end": new_end,
    }



def _api_check_update_booking_conflicts(db, event_id, existing_status, new_start, new_end):
    if new_start is None or new_end is None:
        return None
    booking_statuses = {"confirmed", "reschedule-confirmed", "reserved", "pending-deposit", "pending"}
    if existing_status not in booking_statuses:
        return None
    blocking = ('confirmed', 'reschedule-confirmed', 'reserved', 'travel', 'admin', 'social')
    conflicts = db.execute_query(
        """
        SELECT client_name FROM bookings
        WHERE id != %s AND status = ANY(%s)
        AND start_time < %s AND end_time > %s
        """,
        (event_id, list(blocking), new_end.isoformat(), new_start.isoformat()),
        fetch=True,
    ) or []
    if conflicts:
        names = [str(row_get(r, "client_name") or "existing booking") for r in conflicts[:2]]
        return jsonify({"error": f"Cannot move to this time — conflicts with: {', '.join(names)}"}), 409
    return None



def _api_apply_update_timing_fields(updates, timing):
    if timing["new_start"] is not None:
        updates["start_time"] = timing["new_start"].isoformat()
        updates["end_time"] = timing["new_end"].isoformat()
    if timing["duration_hours"] is not None:
        updates["duration"] = int(float(timing["duration_hours"]) * 60)



def _api_apply_update_identity_fields(updates, data):
    if "client_name" in data or "title" in data:
        updates["client_name"] = data.get("client_name") or data.get("title")
    if "client_phone" in data:
        updates["phone"] = data.get("client_phone")
    if "location_type" in data:
        updates["type"] = data.get("location_type")
    if "experience_type" in data:
        updates["experience"] = data.get("experience_type")
    if "organise_other_escort" in data:
        oe = _normalize_organise_other_escort(data.get("organise_other_escort"))
        updates["organise_other_escort"] = True if oe == "yes" else (False if oe == "no" else None)
    if "notes" in data:
        updates["special_requests"] = data.get("notes")
        updates["notes"] = data.get("notes")
    if "deposit_reference" in data:
        updates["deposit_reference"] = str(data.get("deposit_reference") or "").strip()



def _api_apply_update_preferences_field(updates, data):
    if "preferences" not in data:
        return
    updates["preferences"] = _api_normalize_preferences(data.get("preferences"))



def _api_apply_update_status_field(updates, data):
    if "booking_status" not in data:
        return
    bs = data.get("booking_status")
    if bs == "confirmed":
        updates["status"] = "confirmed"
    elif bs == "pending_deposit":
        updates["status"] = "pending-deposit"
    elif bs == "peacock":
        updates["status"] = "reserved"



def _api_apply_update_financial_fields(updates, data):
    if "deposit_amount" in data:
        dep_amount_val = _api_parse_optional_float(data.get("deposit_amount", "none"), ("none", None, ""))
        updates["deposit_amount"] = dep_amount_val
        if dep_amount_val:
            updates["deposit_status"] = "paid"
        else:
            updates["deposit_status"] = "not_required"
    if "total_cost" in data:
        total_cost = data.get("total_cost")
        price_total = _api_parse_optional_float(total_cost)
        if total_cost in (None, "") or price_total is not None:
            updates["price_total"] = price_total
    if "remaining_balance" in data:
        remaining_balance = data.get("remaining_balance")
        remaining_amount = _api_parse_optional_float(remaining_balance)
        if remaining_balance in (None, "") or remaining_amount is not None:
            updates["remaining_amount"] = remaining_amount



def _api_calculate_update_total(booking_row):
    try:
        from templates.confirmations import calculate_price
        dur_raw = row_get(booking_row, "duration", None)
        dur_min = int(dur_raw) if isinstance(dur_raw, (int, float)) else 60
        exp = row_get(booking_row, "experience", None)
        loc = (row_get(booking_row, "type", None) or "incall").lower()
        c = int(calculate_price(dur_min, experience_type=exp, incall_outcall=loc))
        if c > 0:
            return float(c)
    except Exception:
        pass
    return None



def _api_recalculate_update_remaining(db, event_id, data, updates):
    if ("deposit_amount" not in data and "total_cost" not in data) or "remaining_balance" in data:
        return
    try:
        cur_brow = db.execute_query(
            "SELECT experience, type, duration, price_total, deposit_amount FROM bookings WHERE id = %s",
            (event_id,), fetch=True,
        )
        if not cur_brow:
            return
        br = cur_brow[0]
        dep_for_rem = updates.get("deposit_amount") if "deposit_amount" in updates else float(row_get(br, "deposit_amount") or 0)
        computed_total = _api_calculate_update_total(br)
        price_for_rem = computed_total or updates.get("price_total") or float(row_get(br, "price_total") or 0) or None
        if price_for_rem and dep_for_rem is not None:
            updates["remaining_amount"] = max(0, price_for_rem - dep_for_rem)
            if computed_total:
                updates["price_total"] = computed_total
    except Exception as _calc_err:
        logger.warning("Could not auto-calculate remaining_amount: %s", _calc_err)



def _api_execute_booking_update(db, event_id, updates):
    set_clauses = ", ".join(f"{k} = %s" for k in updates)
    values = list(updates.values()) + [event_id]
    db.execute_query(
        f"UPDATE bookings SET {set_clauses}, updated_at = NOW() WHERE id = %s",
        values,
    )



def _api_send_update_transition_push(db, event_id, existing_status, next_status):
    if existing_status in {"pending-deposit", "pending"} and next_status in {"confirmed", "reschedule-confirmed"}:
        try:
            sent = send_new_booking_push_for_booking_id(db, event_id)
            if sent > 0:
                logger.info("Sent %s push notification(s) for graphite-to-basil transition booking %s", sent, event_id)
        except Exception as push_err:
            logger.warning("Status-transition push notification failed for %s: %s", event_id, push_err)
        return
    if existing_status == "reserved" and next_status in {"confirmed", "reschedule-confirmed"}:
        try:
            sent = send_deposit_paid_push_for_booking_id(db, event_id)
            if sent > 0:
                logger.info("Sent %s push notification(s) for peacock-to-basil transition booking %s", sent, event_id)
        except Exception as push_err:
            logger.warning("Status-transition push notification failed for %s: %s", event_id, push_err)



@schedule_bp.route("/schedule/api/events/<event_id>", methods=["PATCH"])
def update_event_api(event_id):
    """API endpoint for updating an existing booking."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503
        row = _api_fetch_booking_for_update(db, event_id)
        if not row:
            return jsonify({"error": "Booking not found"}), 404

        data = request.get_json() or {}
        timing = _api_prepare_update_timing(row, data, _get_local_timezone())
        conflict_response = _api_check_update_booking_conflicts(
            db,
            event_id,
            timing["existing_status"],
            timing["new_start"],
            timing["new_end"],
        )
        if conflict_response:
            return conflict_response

        updates = {}
        _api_apply_update_timing_fields(updates, timing)
        _api_apply_update_identity_fields(updates, data)
        _api_apply_update_preferences_field(updates, data)
        _api_apply_update_status_field(updates, data)
        _api_apply_update_financial_fields(updates, data)
        _api_recalculate_update_remaining(db, event_id, data, updates)
        if not updates:
            return jsonify({"success": True, "event_id": event_id, "message": "No changes"}), 200

        _api_execute_booking_update(db, event_id, updates)
        _mobile_sync_cache_invalidate()
        next_status = str(updates.get("status", timing["existing_status"]) or "").strip().lower()
        _api_send_update_transition_push(db, event_id, timing["existing_status"], next_status)
        return jsonify({
            "success": True,
            "event_id": event_id,
            "message": "Event updated successfully"
        })

    except Exception as e:
        logger.error(f"Error updating event: {e}")
        return jsonify({"error": str(e)}), 500

@schedule_bp.route("/schedule/api/events/<event_id>", methods=["DELETE"])
def delete_event_api(event_id):
    """API endpoint for deleting a booking.

    Query parameters:
    - notify_client: 'true' to send cancellation SMS to client, 'false' (default) for silent deletion
    """
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    notify_client = request.args.get('notify_client', 'false').lower() == 'true'

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503

        # Fetch booking details before deletion
        rows = db.execute_query(
            "SELECT phone, client_name, start_time FROM bookings WHERE id = %s",
            (event_id,),
            fetch=True,
        ) or []
        phone_number = None
        start_dt = None
        if rows:
            phone_number = str(row_get(rows[0], "phone") or "") or None
            start_dt = row_get(rows[0], "start_time")

        # Delete associated travel blocks via conversation_states
        if phone_number:
            try:
                travel_row = db.execute_query(
                    "SELECT travel_outbound_event_id, travel_return_event_id FROM conversation_states WHERE phone_number = %s",
                    (phone_number,),
                    fetch=True,
                )
                if travel_row:
                    outbound_id = row_get(travel_row[0], 'travel_outbound_event_id', None)
                    return_id = row_get(travel_row[0], 'travel_return_event_id', None)
                    _delete_travel_time_blocks(db, outbound_id, return_id)
                    if outbound_id or return_id:
                        db.execute_query(
                            "UPDATE conversation_states SET travel_outbound_event_id = NULL, travel_return_event_id = NULL WHERE phone_number = %s",
                            (phone_number,),
                        )
            except Exception as e:
                logger.warning(f"Could not delete travel blocks for event {event_id}: {e}")

        db.execute_query("DELETE FROM bookings WHERE id = %s", (event_id,))

        if notify_client and phone_number:
            try:
                from services.sms_service import send_sms
                tz = _get_local_timezone()
                if start_dt:
                    start_local = start_dt.astimezone(tz) if getattr(start_dt, "tzinfo", None) else tz.localize(start_dt)
                    time_str = start_local.strftime("%A %d/%m/%Y %I:%M%p")
                else:
                    time_str = "your scheduled time"
                send_sms(phone_number, f"Your booking on {time_str} has been cancelled. We're sorry for any inconvenience.")
                logger.info(f"Sent cancellation SMS to {phone_number} for event {event_id}")
            except Exception as e:
                logger.warning(f"Could not send cancellation SMS: {e}")

        _mobile_sync_cache_invalidate()
        return jsonify({
            "success": True,
            "event_id": event_id,
            "message": "Event deleted" + (" and client notified" if notify_client else " silently without notifying client")
        })

    except Exception as e:
        logger.error(f"Error deleting event: {e}")
        return jsonify({"error": str(e)}), 500


@schedule_bp.route("/schedule/api/events/<event_id>/cancel-preview", methods=["GET"])
def cancel_event_preview(event_id):
    """Return the SMS message that would be sent for a cancellation, without cancelling."""
    if not _is_schedule_authenticated():
        return jsonify({"error": "Unauthorized"}), 401
    try:
        from utils.row_utils import row_get as _rg
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503
        rows = db.execute_query(
            "SELECT phone, client_name, start_time, deposit_status, deposit_amount FROM bookings WHERE id = %s",
            (event_id,), fetch=True,
        ) or []
        if not rows:
            return jsonify({"error": "Booking not found"}), 404

        phone = _rg(rows[0], "phone") or ""
        client_name = _rg(rows[0], "client_name") or "there"
        start_dt = _rg(rows[0], "start_time")
        dep_status = str(_rg(rows[0], "deposit_status") or "")
        dep_amount = _rg(rows[0], "deposit_amount") or 0

        state_row = db.execute_query(
            "SELECT deposit_paid FROM conversation_states WHERE phone_number = %s",
            (phone,), fetch=True,
        )
        deposit_paid = bool(state_row and _rg(state_row[0], "deposit_paid", False))
        if not deposit_paid:
            deposit_paid = dep_status == "paid"
        try:
            deposit_amount = int(float(dep_amount))
        except (TypeError, ValueError):
            deposit_amount = 0

        tz = _get_local_timezone()
        if start_dt:
            start_local = start_dt.astimezone(tz) if getattr(start_dt, "tzinfo", None) else tz.localize(start_dt)
            start_str = start_local.strftime("%A %d/%m/%Y %I:%M%p")
        else:
            start_str = "your scheduled time"

        webform_url = f"{config.get_base_url()}/booking"
        try:
            from core.webform_security import get_webform_url
            webform_url = get_webform_url(phone)
        except Exception:
            pass

        escort_name = config.get_escort_name()
        if deposit_paid and deposit_amount:
            msg = (
                f"Hi {client_name} I'm very sorry but I need to cancel your booking scheduled for {start_str}. "
                f"I apologise for any inconvenience. In order to issue you a refund for your deposit of ${deposit_amount} "
                f"please forward your banking details so I can process you a full refund. "
                f"If you'd like to rebook for another time please text me back or instead fill in my booking webform {webform_url} "
                f"Hope to see you soon {escort_name}"
            )
        else:
            msg = (
                f"Hi {client_name} I'm very sorry but I need to cancel your booking scheduled for {start_str}. "
                f"I apologise for any inconvenience. If you'd like to rebook for another time please text me back or instead fill in my booking webform {webform_url} "
                f"Hope to see you soon {escort_name}"
            )
        return jsonify({"message": msg, "phone": phone})
    except Exception as e:
        logger.error("cancel preview error: %s", e)
        return jsonify({"error": str(e)}), 500


@schedule_bp.route("/schedule/api/events/<event_id>/cancel", methods=["POST"])
def cancel_event_like_schedule_page_api(event_id):
    """Cancel a booking using the same workflow as the schedule page."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    fake_request = SimpleNamespace(form=MultiDict({"event_id": event_id}))
    result = _handle_cancellation(fake_request)
    if result.get("success"):
        _mobile_sync_cache_invalidate()
    return jsonify(result), (200 if result.get("success") else 400)


@schedule_bp.route("/schedule/api/admin/backfill-financial-fields", methods=["POST"])
def backfill_financial_fields():
    """Temporary admin endpoint to backfill financial fields for existing bookings."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"error": "Database unavailable"}), 503

        results = {}

        # Update bookings where price_total is NULL but deposit_amount exists
        db.execute_query("""
            UPDATE bookings
            SET price_total = deposit_amount * 2
            WHERE price_total IS NULL
              AND deposit_amount IS NOT NULL
              AND deposit_amount > 0
        """)
        results["price_total_from_deposit"] = db.cursor.rowcount

        # Update bookings where remaining_amount is NULL but price_total and deposit_amount exist
        db.execute_query("""
            UPDATE bookings
            SET remaining_amount = price_total - deposit_amount
            WHERE remaining_amount IS NULL
              AND price_total IS NOT NULL
              AND deposit_amount IS NOT NULL
        """)
        results["remaining_amount"] = db.cursor.rowcount

        # Set deposit_reference to empty string if NULL
        db.execute_query("""
            UPDATE bookings
            SET deposit_reference = ''
            WHERE deposit_reference IS NULL
        """)
        results["deposit_reference"] = db.cursor.rowcount

        # For bookings with no deposit (reserved/peacock), set a default price_total
        db.execute_query("""
            UPDATE bookings
            SET price_total = 600,
                remaining_amount = 600
            WHERE price_total IS NULL
              AND (deposit_status = 'not_required' OR deposit_amount IS NULL OR deposit_amount = 0)
              AND status IN ('reserved', 'confirmed', 'reschedule-confirmed')
        """)
        results["default_price_no_deposit"] = db.cursor.rowcount

        db.conn.commit()
        _mobile_sync_cache_invalidate()

        return jsonify({
            "success": True,
            "results": results,
            "message": "Financial fields backfilled successfully"
        }), 200

    except Exception as e:
        logger.error(f"Error backfilling financial fields: {e}")
        return jsonify({"error": str(e)}), 500


@schedule_bp.route("/schedule/api/events/<event_id>/reschedule", methods=["POST"])
def reschedule_event_like_schedule_page_api(event_id):
    """Send the same reschedule request flow used by the schedule page."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json() or {}
    form = MultiDict(
        {
            "event_id": event_id,
            "phone_number": (data.get("phone_number") or "").strip(),
            "client_name": (data.get("client_name") or "Client").strip(),
            "original_time": (data.get("original_time") or "").strip(),
            "new_date": (data.get("new_date") or "").strip(),
            "new_time": (data.get("new_time") or "").strip(),
            "custom_message": (data.get("custom_message") or "").strip(),
        }
    )
    fake_request = SimpleNamespace(form=form)
    result = _handle_reschedule(fake_request)
    if result.get("success"):
        _mobile_sync_cache_invalidate()
    status_code = 200 if result.get("success") else (409 if "conflicts with" in str(result.get("error", "")).lower() else 400)
    return jsonify(result), status_code


@schedule_bp.route('/schedule/api/clear-events', methods=['POST'])
def clear_events_api():
    """Delete all bookings for a given period (day / month / year). Password protected."""
    data = request.get_json() or {}
    password = data.get('password', '')
    period = data.get('period', '')
    date_str = data.get('date', '')

    if not verify_password(password):
        return jsonify({'success': False, 'error': 'Incorrect password'}), 403

    if period not in ('day', 'month', 'year'):
        return jsonify({'success': False, 'error': 'Invalid period'}), 400

    try:
        from datetime import date as _date
        base_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else _date.today()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date'}), 400

    from config import get_effective_escort_timezone
    tz = pytz.timezone(get_effective_escort_timezone())

    if period == 'day':
        start_dt = tz.localize(datetime.combine(base_date, datetime.min.time()))
        end_dt = start_dt + timedelta(days=1)
        label = base_date.strftime('%d %b %Y')
    elif period == 'month':
        import calendar as _cal
        start_dt = tz.localize(datetime(base_date.year, base_date.month, 1))
        last_day = _cal.monthrange(base_date.year, base_date.month)[1]
        end_dt = tz.localize(datetime(base_date.year, base_date.month, last_day, 23, 59, 59))
        label = base_date.strftime('%B %Y')
    else:  # year
        start_dt = tz.localize(datetime(base_date.year, 1, 1))
        end_dt = tz.localize(datetime(base_date.year, 12, 31, 23, 59, 59))
        label = str(base_date.year)

    try:
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return jsonify({"success": False, "error": "Database unavailable"}), 503
        count_rows = db.execute_query(
            "SELECT COUNT(*) AS cnt FROM bookings WHERE start_time >= %s AND start_time < %s",
            (start_dt.isoformat(), end_dt.isoformat()),
            fetch=True,
        ) or []
        deleted = int(row_get(count_rows[0], "cnt") or 0) if count_rows else 0
        db.execute_query(
            "DELETE FROM bookings WHERE start_time >= %s AND start_time < %s",
            (start_dt.isoformat(), end_dt.isoformat()),
        )
        logger.info("Clear schedule: deleted %d bookings for %s (%s)", deleted, period, label)
        if deleted:
            _mobile_sync_cache_invalidate()
        return jsonify({'success': True, 'message': f'Deleted {deleted} booking(s) for {label}.'})
    except Exception as e:
        logger.error("clear_events_api error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


def _api_fetch_export_rows(db, tz_name, date_str):
    return db.execute_query(
        """
        SELECT id, start_time, end_time, client_name, experience, type,
               status, price_total, deposit_status, deposit_amount,
               special_requests, notes
        FROM bookings
        WHERE DATE(start_time AT TIME ZONE %s) = %s
        ORDER BY start_time ASC
        """,
        (tz_name, date_str),
        fetch=True,
    ) or []



def _api_build_export_styles(PatternFill, Font, Side, Border):
    thin = Side(style='thin', color='444444')
    return {
        'header_fill': PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid'),
        'header_font': Font(bold=True, color='FFEB3B', size=11),
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'row_fill_light': PatternFill(start_color='16213e', end_color='16213e', fill_type='solid'),
        'row_fill_alt': PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid'),
        'row_font': Font(color='FFFFFF', size=10),
    }



def _api_initialize_export_sheet(ws, date_obj, headers, styles, Alignment, openpyxl):
    ws.title = date_obj.strftime('%d %b %Y')
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['border']
    ws.row_dimensions[1].height = 20
    col_widths = [10, 10, 12, 20, 20, 20, 18, 10, 36]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w



def _api_build_export_row_data(row, tz):
    start = row_get(row, 'start_time')
    end = row_get(row, 'end_time')
    status = str(row_get(row, 'status') or 'reserved')
    price_total = row_get(row, 'price_total')
    dep_status = str(row_get(row, 'deposit_status') or 'not_required')
    dep_amount = float(row_get(row, 'deposit_amount') or 0)
    start_local = start.astimezone(tz) if getattr(start, 'tzinfo', None) else tz.localize(start)
    end_local = end.astimezone(tz) if getattr(end, 'tzinfo', None) else tz.localize(end)
    duration_mins = int((end_local - start_local).total_seconds() / 60)
    duration_str = f"{duration_mins // 60}h {duration_mins % 60}m" if duration_mins % 60 else f"{duration_mins // 60}h"
    price = str(float(price_total)) if price_total is not None else (
        str(dep_amount) if dep_status == 'paid' else 'N/A'
    )
    special_requests = str(row_get(row, 'special_requests') or row_get(row, 'notes') or '').strip()
    return [
        start_local.strftime('%I:%M %p'),
        end_local.strftime('%I:%M %p'),
        duration_str,
        str(row_get(row, 'client_name') or 'N/A'),
        str(row_get(row, 'experience') or 'N/A'),
        str(row_get(row, 'type') or 'N/A'),
        status.replace('-', ' ').title(),
        price,
        special_requests,
    ]



def _api_write_export_rows(ws, rows, tz, styles, Alignment):
    write_row = 2
    for row in rows:
        try:
            row_data = _api_build_export_row_data(row, tz)
            fill = styles['row_fill_light'] if write_row % 2 == 0 else styles['row_fill_alt']
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=write_row, column=col, value=val)
                cell.font = styles['row_font']
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = styles['border']
            write_row += 1
        except Exception as row_err:
            logger.warning('Skipping malformed booking row %s in export: %s', row_get(row, 'id', '?'), row_err)



def _api_add_export_title(ws, date_obj, headers, Font, PatternFill, Alignment):
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"Schedule — {date_obj.strftime('%A %d %B %Y')}")
    title_cell.font = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill = PatternFill(start_color='0d0d1a', end_color='0d0d1a', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 24



def _api_build_export_response(wb, io, Response, date_str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'schedule_{date_str}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )



@schedule_bp.route('/schedule/api/export-excel', methods=['GET'])
def export_excel_api():
    """Export bookings for a given day to an Excel (.xlsx) file."""
    authenticated = _is_schedule_authenticated()
    if not authenticated:
        from flask import redirect, url_for
        return redirect(url_for('schedule.schedule_management'))

    from utils.timezone import get_current_datetime

    date_str = request.args.get('date', get_current_datetime().strftime('%Y-%m-%d'))
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return "Invalid date", 400

    try:
        import io

        import openpyxl
        from flask import Response
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        tz = _get_local_timezone()
        date_obj = tz.localize(datetime.strptime(date_str, '%Y-%m-%d'))
        db = get_shared_db(config.DATABASE_URL)
        if db is None:
            return "Database unavailable", 503
        rows = _api_fetch_export_rows(db, str(tz), date_str)

        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Start', 'End', 'Duration', 'Client', 'Service', 'Location', 'Status', 'Price', 'Special requests']
        styles = _api_build_export_styles(PatternFill, Font, Side, Border)
        _api_initialize_export_sheet(ws, date_obj, headers, styles, Alignment, openpyxl)
        _api_write_export_rows(ws, rows, tz, styles, Alignment)
        _api_add_export_title(ws, date_obj, headers, Font, PatternFill, Alignment)
        return _api_build_export_response(wb, io, Response, date_str)

    except ImportError:
        return "openpyxl is not installed on the server.", 500
    except Exception as e:
        logger.error(f"export_excel_api error: {e}")
        return f"Export failed: {e}", 500

@schedule_bp.route('/schedule/logout', methods=['GET', 'POST'])
def schedule_logout():
    """Logout from schedule page."""
    session.pop("schedule_authenticated", None)
    from flask import redirect, url_for
    return redirect(url_for('schedule.schedule_management'))


